import json
from functools import wraps
from pathlib import Path

import sentry_sdk
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.mail import EmailMessage
from django.http import Http404
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.shortcuts import redirect
from django.shortcuts import render
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils.exceptions import IllegalCharacterError

from .forms import AnalyseIAForm
from .helpers import normalise_titre_esrs
from .helpers import normalise_titre_pour_nom_de_fichier
from .helpers import synthese_analyse
from .models import AnalyseIA
from api import analyse_ia
from api.exceptions import APIError
from entreprises.decorators import entreprise_qualifiee_requise
from habilitations.models import Habilitation
from reglementations.enums import ESRS
from reglementations.views import tableau_de_bord_menu_context
from utils.xlsx import xlsx_response


def _contexte_analyses(entreprise):
    context = tableau_de_bord_menu_context(entreprise)
    context |= {
        "form": AnalyseIAForm(),
        "analyses_ia": entreprise.analyses_ia.all(),
        "synthese": synthese_analyse(entreprise.analyses_ia.reussies()),
    }
    return context


@login_required
@entreprise_qualifiee_requise
def analyses(request, entreprise_qualifiee):
    return render(
        request, "analyseia/accueil.html", _contexte_analyses(entreprise_qualifiee)
    )


@login_required
@entreprise_qualifiee_requise
@require_http_methods(["POST"])
def ajout_document(request, entreprise_qualifiee):
    data = {**request.POST}
    form = AnalyseIAForm(data=data, files=request.FILES)
    if form.is_valid():
        # les analyses IA étant désormais génériques,
        # on effectue le rattachement à l'entreprise
        form.save()
        entreprise_qualifiee.analyses_ia.add(form.instance)
        messages.success(request, "Document ajouté")
        return redirect("analyseia:analyses", siren=entreprise_qualifiee.siren)
    else:
        return render(
            request,
            "analyseia/accueil.html",
            _contexte_analyses(entreprise_qualifiee),
            status=400,
        )


def analyse_requise(function):
    @wraps(function)
    def wrap(request, id_analyse, *args, **kwargs):
        analyse = get_object_or_404(AnalyseIA, id=id_analyse)

        if not Habilitation.existe(analyse.entreprise, request.user):
            raise PermissionDenied()
        return function(request, analyse, *args, **kwargs)

    return wrap


@login_required
@analyse_requise
@require_http_methods(["POST"])
def suppression(request, analyse):
    entreprise = analyse.entreprise
    analyse.delete()
    analyse.fichier.delete(save=False)
    messages.success(request, "Document supprimé")
    return redirect("analyseia:analyses", siren=entreprise.siren)


@login_required
@analyse_requise
@require_http_methods(["POST"])
def lancement_analyse(request, analyse):
    if analyse.etat != "success":
        try:
            callback_url = request.build_absolute_uri(
                reverse(
                    "analyseia:actualisation_etat",
                    kwargs={
                        "id_analyse": analyse.id,
                    },
                )
            )
            etat = analyse_ia.lancement_analyse(
                analyse.id, analyse.fichier.url, callback_url
            )
            analyse.etat = etat
            analyse.save()
            messages.success(
                request,
                "L'analyse a bien été lancée. Celle-ci peut prendre entre quelques secondes et quelques heures. Vous recevrez un email lorsque les résultats seront disponibles.",
            )
        except APIError as exception:
            messages.error(request, exception)

    return redirect("analyseia:analyses")


@csrf_exempt
@require_http_methods(["POST"])
def actualisation_etat(request, id_analyse):
    # callback pour l'API IA qui renvoie le statut de l'analyse IA à la fin de chaque étape de l'analyse
    analyse = get_object_or_404(AnalyseIA, pk=id_analyse)
    status = request.POST.get("status")
    analyse.etat = status
    if message := request.POST.get("msg"):
        analyse.message = message
    if status == "success":
        analyse.resultat_json = request.POST["resultat_json"]
    analyse.save()
    if status in ("success", "error"):
        path = reverse("analyseia:analyses", kwargs={"siren": analyse.entreprise.siren})
        try:
            _envoie_resultat_ia_email(
                analyse.entreprise,
                f"{request.build_absolute_uri(path)}#onglets",
            )
        except Exception as e:
            with sentry_sdk.new_scope() as scope:
                scope.set_level("info")
                sentry_sdk.capture_exception(e)

    return HttpResponse("OK")


@login_required
@analyse_requise
def resultat(request, analyse, rendu):
    chemin_xlsx = Path(
        settings.BASE_DIR, f"analyseia/xlsx/{rendu}/template_synthese_ESG.xlsx"
    )
    workbook = load_workbook(chemin_xlsx)
    worksheet = workbook[">>>"]
    worksheet["C14"] = ""
    if rendu == "theme":
        worksheet = workbook["Phrases relatives aux ESG"]
    else:
        worksheet = workbook["Phrases relatives aux ESRS"]

    prefixe_ESRS = rendu == "esrs"
    _ajoute_lignes_resultat_ia(worksheet, analyse, True, None, prefixe_ESRS)
    return xlsx_response(workbook, "resultats.xlsx")


# TODO: externaliser ?


def _ajoute_lignes_resultat_ia(
    worksheet, document, avec_nom_fichier, contrainte_esrs, prefixe_ESRS=False
):
    if not document.resultat_json:
        return
    data = json.loads(document.resultat_json)
    for esrs, contenus in data.items():
        for contenu in contenus:
            if esrs != "Non ESRS" and (
                (not contrainte_esrs) or contrainte_esrs in esrs
            ):
                if avec_nom_fichier:
                    ligne = [
                        normalise_titre_esrs(esrs, prefixe_ESRS=prefixe_ESRS),
                        document.nom,
                        contenu["PAGES"],
                        contenu["TEXTS"],
                    ]
                else:
                    ligne = [
                        normalise_titre_esrs(esrs, prefixe_ESRS=prefixe_ESRS),
                        contenu["PAGES"],
                        contenu["TEXTS"],
                    ]
                try:
                    worksheet.append(ligne)
                except IllegalCharacterError:
                    ligne[-1] = ILLEGAL_CHARACTERS_RE.sub("", contenu["TEXTS"])
                    worksheet.append(ligne)


def _envoie_resultat_ia_email(entreprise, resultat_ia_url):
    destinataires = [utilisateur.email for utilisateur in entreprise.users.all()]

    email = EmailMessage(
        to=destinataires,
        from_email=settings.DEFAULT_FROM_EMAIL,
    )
    email.template_id = settings.BREVO_RESULTAT_ANALYSE_IA_TEMPLATE

    email.merge_global_data = {
        "resultat_ia_url": resultat_ia_url,
    }
    email.send()


@login_required
@entreprise_qualifiee_requise
def synthese_resultat(request, entreprise_qualifiee, csrd_id=None):
    rendu = "esrs" if csrd_id else "theme"
    chemin_xlsx = Path(
        settings.BASE_DIR, f"analyseia/xlsx/{rendu}/template_synthese_ESG.xlsx"
    )
    workbook = load_workbook(chemin_xlsx)
    if rendu == "theme":
        worksheet = workbook["Phrases relatives aux ESG"]
        documents = entreprise_qualifiee.analyses_ia.reussies()
    else:
        worksheet = workbook["Phrases relatives aux ESRS"]
        from reglementations.models import RapportCSRD

        csrd = get_object_or_404(RapportCSRD, id=csrd_id)
        documents = csrd.documents_analyses

    prefixe_ESRS = rendu == "esrs"
    for document in documents:
        _ajoute_lignes_resultat_ia(worksheet, document, True, None, prefixe_ESRS)
    return xlsx_response(workbook, "synthese_resultats.xlsx")


@login_required
@entreprise_qualifiee_requise
def synthese_resultat_par_ESRS(request, entreprise_qualifiee, code_esrs, csrd_id=None):
    if code_esrs not in ESRS.codes():
        raise Http404

    rendu = "esrs" if csrd_id else "theme"
    prefixe_ESRS = rendu == "esrs"
    chemin_xlsx = Path(
        settings.BASE_DIR,
        f"analyseia/xlsx/{rendu}/template_synthese_{code_esrs[0]}.xlsx",
    )
    workbook = load_workbook(chemin_xlsx)
    worksheet = workbook[">>>"]
    titre = normalise_titre_esrs(f"ESRS {code_esrs}", prefixe_ESRS=prefixe_ESRS)
    worksheet["C14"] = titre
    if rendu == "theme":
        worksheet = workbook["Phrases relatives aux ESG"]
        documents = entreprise_qualifiee.analyses_ia.reussies()
    else:
        worksheet = workbook["Phrases relatives aux ESRS"]
        from reglementations.models import RapportCSRD

        csrd = get_object_or_404(RapportCSRD, id=csrd_id)
        documents = csrd.documents_analyses

    for document in documents:
        _ajoute_lignes_resultat_ia(worksheet, document, True, code_esrs, prefixe_ESRS)
    if rendu == "theme":
        titre_pour_nom_de_fichier = normalise_titre_pour_nom_de_fichier(titre)
        nom_de_fichier = f"resultats_{titre_pour_nom_de_fichier}.xlsx"
    else:
        nom_de_fichier = f"resultats_ESRS_{code_esrs}.xlsx"
    return xlsx_response(workbook, nom_de_fichier)


# Fragments / HTMX


@login_required
def statut_analyse_ia(request, id_analyse):
    analyse = get_object_or_404(AnalyseIA, pk=id_analyse)
    return render(
        request,
        "fragments/statut_analyse.html",
        {
            "analyse": analyse,
        },
    )

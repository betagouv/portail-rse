from io import BytesIO

import pytest
from django.urls import reverse
from openpyxl import load_workbook


@pytest.mark.parametrize("rendu", ["theme", "esrs"])
def test_telechargement_des_resultats_IA_d_un_document_au_format_xlsx(
    rendu, client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    document = entreprise.analyses_ia.create(
        etat="success",
        resultat_json="""{
  "ESRS E1": [
    {
      "PAGES": 1,
      "TEXTS": "A"
    }
  ],
  "ESRS E2 : Pollution": [
    {
      "PAGES": 6,
      "TEXTS": "B"
    },
    {
      "PAGES": 7,
      "TEXTS": "C"
    }
  ]
  }""",
    )

    client.force_login(alice)

    response = client.get(
        reverse("analyseia:resultat", args=[document.id, rendu]),
    )

    assert response["Content-Disposition"] == "filename=resultats.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook[">>>"]
    assert not onglet["C14"].value
    if rendu == "esrs":
        onglet = workbook["Phrases relatives aux ESRS"]
        assert onglet["A1"].value == "ESRS"
    else:
        onglet = workbook["Phrases relatives aux ESG"]
        assert onglet["A1"].value == "Thème"
    assert onglet["B1"].value == "Fichier"
    assert onglet["C1"].value == "Page"
    assert onglet["D1"].value == "Phrase"
    if rendu == "esrs":
        assert onglet["A2"].value == "ESRS E1 - Changement climatique"
    else:
        assert onglet["A2"].value == "Changement climatique"
    assert onglet["C2"].value == 1
    assert onglet["D2"].value == "A"
    if rendu == "esrs":
        assert onglet["A3"].value == "ESRS E2 - Pollution"
    else:
        assert onglet["A3"].value == "Pollution"
    assert onglet["C3"].value == 6
    assert onglet["D3"].value == "B"
    if rendu == "esrs":
        assert onglet["A3"].value == "ESRS E2 - Pollution"
    else:
        assert onglet["A4"].value == "Pollution"
    assert onglet["C4"].value == 7
    assert onglet["D4"].value == "C"


@pytest.mark.parametrize("rendu", ["theme", "esrs"])
def test_telechargement_des_resultats_IA_d_une_analyse_non_terminee(
    rendu, client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    analyse = entreprise.analyses_ia.create(
        etat="processing",
    )

    client.force_login(alice)

    response = client.get(
        reverse("analyseia:resultat", args=[analyse.id, rendu]),
    )

    assert response["Content-Disposition"] == "filename=resultats.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook[">>>"]
    assert not onglet["A2"].value
    assert not onglet["B2"].value
    assert not onglet["C2"].value
    assert not onglet["D2"].value


@pytest.mark.parametrize("rendu", ["theme", "esrs"])
def test_telechargement_des_resultats_IA_d_un_document_inexistant(
    rendu, client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    client.force_login(alice)

    response = client.get(
        reverse("analyseia:resultat", args=[42, rendu]),
    )

    assert response.status_code == 404


@pytest.mark.parametrize("rendu", ["theme", "esrs"])
def test_telechargement_des_resultats_IA_d_un_document_redirige_vers_la_connexion_si_non_connecté(
    rendu, client, entreprise_factory, alice, analyse
):
    entreprise = entreprise_factory(siren="000000089", utilisateur=alice)

    response = client.get(
        reverse("analyseia:resultat", args=[analyse.id, rendu]),
    )

    assert response.status_code == 302


def test_telechargement_des_resultats_ia_de_l_ensemble_des_documents_au_format_xlsx_lié_à_une_entreprise(
    client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice, siren="123456789")
    entreprise.analyses_ia.create(
        etat="success",
        resultat_json="""{
  "ESRS E1": [
    {
      "PAGES": 1,
      "TEXTS": "A"
    }
   ],
    "Non ESRS": [
    {
      "PAGES": 22,
      "TEXTS": "X"
    }
  ]
  }""",
    )
    entreprise.analyses_ia.create(
        etat="success",
        resultat_json="""{
  "ESRS E2": [
    {
      "PAGES": 6,
      "TEXTS": "B"
    }
  ]
  }""",
    )
    entreprise.analyses_ia.create(
        etat="processing",
    )
    client.force_login(alice)

    response = client.get(
        reverse("analyseia:synthese_resultat", args=[entreprise.siren]),
    )

    assert response["Content-Disposition"] == "filename=synthese_resultats.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook[">>>"]
    assert onglet["C14"].value == "Synthèse générique"
    onglet = workbook["Phrases relatives aux ESG"]
    assert onglet["A1"].value == "Thème"
    assert onglet["B1"].value == "Fichier"
    assert onglet["C1"].value == "Page"
    assert onglet["D1"].value == "Phrase"
    assert onglet["A2"].value == "Changement climatique"
    assert onglet["C2"].value == 1
    assert onglet["D2"].value == "A"
    assert onglet["A3"].value == "Pollution"
    assert onglet["C3"].value == 6
    assert onglet["D3"].value == "B"


def test_telechargement_des_resultats_ia_de_l_ensemble_des_documents_au_format_xlsx_lié_à_une_CSRD(
    client, csrd, alice
):
    csrd.analyses_ia.create(
        etat="success",
        resultat_json="""{
  "ESRS E1": [
    {
      "PAGES": 1,
      "TEXTS": "A"
    }
   ],
    "Non ESRS": [
    {
      "PAGES": 22,
      "TEXTS": "X"
    }
  ]
  }""",
    )
    csrd.analyses_ia.create(
        etat="success",
        resultat_json="""{
  "ESRS E2": [
    {
      "PAGES": 6,
      "TEXTS": "B"
    }
  ]
  }""",
    )
    csrd.analyses_ia.create(
        etat="processing",
    )
    client.force_login(alice)

    response = client.get(
        reverse("analyseia:synthese_resultat", args=[csrd.entreprise.siren, csrd.id]),
    )

    assert response["Content-Disposition"] == "filename=synthese_resultats.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook[">>>"]
    assert onglet["C14"].value == "Synthèse générique"
    onglet = workbook["Phrases relatives aux ESRS"]
    assert onglet["A1"].value == "ESRS"
    assert onglet["B1"].value == "Fichier"
    assert onglet["C1"].value == "Page"
    assert onglet["D1"].value == "Phrase"
    assert onglet["A2"].value == "ESRS E1 - Changement climatique"
    assert onglet["C2"].value == 1
    assert onglet["D2"].value == "A"
    assert onglet["A3"].value == "ESRS E2 - Pollution"
    assert onglet["C3"].value == 6
    assert onglet["D3"].value == "B"


def test_telechargement_des_resultats_IA_de_l_ensemble_des_documents_d_un_rapport_csrd_inexistant(
    client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    client.force_login(alice)

    response = client.get(
        reverse("analyseia:synthese_resultat", args=[entreprise.siren, 42]),
    )

    assert response.status_code == 404


def test_telechargement_des_resultats_IA_de_l_ensemble_des_documents_redirige_vers_la_connexion_si_non_connecté(
    client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)

    response = client.get(
        reverse("analyseia:synthese_resultat", args=[entreprise.siren]),
    )

    assert response.status_code == 302

    response = client.get(
        reverse("analyseia:synthese_resultat", args=[entreprise.siren, 42]),
    )

    assert response.status_code == 302


def test_telechargement_des_resultats_par_ESRS_au_format_xlsx_lié_à_une_entreprise(
    client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    entreprise.analyses_ia.create(
        etat="success",
        resultat_json="""{
  "ESRS E1": [
    {
      "PAGES": 1,
      "TEXTS": "A"
    }
  ],
  "ESRS E5": [
    {
      "PAGES": 4,
      "TEXTS": "B"
    }
  ],
  "Non ESRS": [
    {
      "PAGES": 22,
      "TEXTS": "X"
    }
  ]
  }""",
    )
    entreprise.analyses_ia.create(
        etat="success",
        resultat_json="""{
  "ESRS E5": [
    {
      "PAGES": 6,
      "TEXTS": "C"
    }
  ]
  }""",
    )
    entreprise.analyses_ia.create(
        etat="processing",
    )
    client.force_login(alice)

    response = client.get(
        reverse("analyseia:synthese_resultat_par_ESRS", args=[entreprise.siren, "E5"]),
    )

    assert (
        response["Content-Disposition"]
        == "filename=resultats_utilisation_des_ressources_et_economie_circulaire.xlsx"
    )
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook[">>>"]
    TITRE_ESRS_E5 = "Utilisation des ressources et économie circulaire"
    assert onglet["C14"].value == TITRE_ESRS_E5
    onglet = workbook["Phrases relatives aux ESG"]
    assert onglet["A1"].value == "Thèmes"
    assert onglet["B1"].value == "Fichier"
    assert onglet["C1"].value == "Page"
    assert onglet["D1"].value == "Phrase"
    assert onglet["A2"].value == TITRE_ESRS_E5
    assert onglet["C2"].value == 4
    assert onglet["D2"].value == "B"
    assert onglet["A3"].value == TITRE_ESRS_E5
    assert onglet["C3"].value == 6
    assert onglet["D3"].value == "C"


def test_telechargement_des_resultats_par_ESRS_au_format_xlsx_lié_à_une_CSRD(
    client, csrd, alice
):
    csrd.analyses_ia.create(
        etat="success",
        resultat_json="""{
  "ESRS E1": [
    {
      "PAGES": 1,
      "TEXTS": "A"
    }
  ],
  "ESRS E2": [
    {
      "PAGES": 4,
      "TEXTS": "B"
    }
  ],
  "Non ESRS": [
    {
      "PAGES": 22,
      "TEXTS": "X"
    }
  ]
  }""",
    )
    csrd.analyses_ia.create(
        etat="success",
        resultat_json="""{
  "ESRS E2": [
    {
      "PAGES": 6,
      "TEXTS": "C"
    }
  ]
  }""",
    )
    csrd.analyses_ia.create(
        etat="processing",
    )
    client.force_login(alice)

    response = client.get(
        reverse(
            "analyseia:synthese_resultat_par_ESRS",
            args=[csrd.entreprise.siren, "E2", csrd.id],
        ),
    )

    assert response["Content-Disposition"] == "filename=resultats_ESRS_E2.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook[">>>"]
    assert onglet["C14"].value == "ESRS E2 - Pollution"
    onglet = workbook["Phrases relatives aux ESRS"]
    assert onglet["A1"].value == "ESRS"
    assert onglet["B1"].value == "Fichier"
    assert onglet["C1"].value == "Page"
    assert onglet["D1"].value == "Phrase"
    assert onglet["A2"].value == "ESRS E2 - Pollution"
    assert onglet["C2"].value == 4
    assert onglet["D2"].value == "B"
    assert onglet["A3"].value == "ESRS E2 - Pollution"
    assert onglet["C3"].value == 6
    assert onglet["D3"].value == "C"


def test_telechargement_des_resultats_IA_par_ESRS_d_une_entreprise_inexistante(
    client, alice
):
    client.force_login(alice)

    response = client.get(
        reverse("analyseia:synthese_resultat_par_ESRS", args=["123456789", "E2"]),
    )

    assert response.status_code == 404

    response = client.get(
        reverse("analyseia:synthese_resultat_par_ESRS", args=["123456789", "E2", 42]),
    )

    assert response.status_code == 404


def test_telechargement_des_resultats_IA_par_ESRS_d_un_ESRS_inexistant(
    client,
    entreprise_factory,
    alice,
    csrd,
):
    entreprise = entreprise_factory(utilisateur=alice, siren="123456789")
    client.force_login(alice)

    response = client.get(
        reverse("analyseia:synthese_resultat_par_ESRS", args=["123456789", "H8"]),
    )

    assert response.status_code == 404

    response = client.get(
        reverse(
            "analyseia:synthese_resultat_par_ESRS", args=["123456789", "H8", csrd.id]
        ),
    )

    assert response.status_code == 404


def test_telechargement_des_resultats_IA_par_ESRS_redirige_vers_la_connexion_si_non_connecté(
    client,
    entreprise_factory,
    alice,
    csrd,
):
    entreprise = entreprise_factory(utilisateur=alice, siren="123456789")

    response = client.get(
        reverse("analyseia:synthese_resultat_par_ESRS", args=[entreprise.siren, "E2"]),
    )

    assert response.status_code == 302

    response = client.get(
        reverse(
            "analyseia:synthese_resultat_par_ESRS",
            args=[entreprise.siren, "E2", csrd.id],
        ),
    )

    assert response.status_code == 302

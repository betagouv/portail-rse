from io import BytesIO

from django.urls import reverse
from openpyxl import load_workbook

from vsme.models import Indicateur
from vsme.models import RapportVSME


def test_telechargement_d_un_rapport_vsme_au_format_xlsx_B1_intégralement_rempli(
    client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    rapport_vsme = RapportVSME.objects.create(entreprise=entreprise, annee=2025)
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-24-a",
        data={"choix_module": "complet"},
    )
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-24-b",
        data={"non_pertinent": False, "omission_informations": ["B6", "B8"]},
    )
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-24-c",
        data={"type_perimetre": "consolidee"},
    )
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-24-d",
        data={
            "filiales": [
                {
                    "denomination_filiale": "filiale 1",
                    "adresse": "adresse 1",
                    "pays": "FRA",
                    "code_postal": "33000",
                    "commentaire": "commentaire F1",
                },
                {
                    "denomination_filiale": "filiale 2",
                    "adresse": "adresse 2",
                    "pays": "DEU",
                    "code_postal": "42000",
                    "commentaire": "commentaire F2",
                },
            ]
        },
    )
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-24-e-i",
        data={"forme_juridique": "54", "coopérative": True},
    )
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-24-e-ii",
        data={"nace": ["01.12", "01.16"]},
    )
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-24-e-iii",
        data={"bilan": 54321},
    )
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-24-e-iv",
        data={"chiffre_affaires": 98765},
    )
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-24-e-v",
        data={"methode_comptabilisation": "ETP", "nombre_salaries": 42.0},
    )
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-24-e-vi",
        data={"pays": ["DEU", "FRA"]},
    )
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-24-e-vii",
        data={
            "sites": [
                {
                    "id_site": 1,
                    "nom_site": "site 1",
                    "adresse": "adresse site 1",
                    "code_postal": "75000",
                    "ville": "Paris",
                    "pays": "FRA",
                    "geolocalisation": "[2.294381,48.858099]",
                },
                {
                    "id_site": 2,
                    "nom_site": "site 2",
                    "adresse": "adresse site 2",
                    "code_postal": "12345",
                    "ville": "Berlin",
                    "pays": "DEU",
                    "geolocalisation": "[3,3]",
                },
            ]
        },
    )
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-25",
        data={
            "non_pertinent": False,
            "certifications": [
                {
                    "nom_certification": "label",
                    "emetteur": "émetteur 1",
                    "date_obtention": "2000-12-12",
                    "score": "10/20",
                    "commentaire": "un commentaire",
                },
                {
                    "nom_certification": "certif",
                    "emetteur": "émetteur 2",
                    "date_obtention": "2020-04-05",
                    "score": "bon",
                    "commentaire": "",
                },
            ],
        },
    )
    client.force_login(alice)

    response = client.get(f"/vsme/{rapport_vsme.id}/export/xlsx")

    assert response["Content-Disposition"] == "filename=vsme.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook["B1"]
    assert onglet["A4"].value == "Module complet"
    assert onglet["B4"].value == "B6 - Eau"
    assert onglet["B5"].value == "B8 - Effectifs : caractéristiques générales"
    assert onglet["C4"].value == "Base consolidée"
    assert onglet["D4"].value == "filiale 1"
    assert onglet["E4"].value == "adresse 1"
    assert onglet["F4"].value == "FRANCE"
    assert onglet["G4"].value == "33000"
    assert onglet["H4"].value == "commentaire F1"
    assert onglet["D5"].value == "filiale 2"
    assert onglet["E5"].value == "adresse 2"
    assert onglet["F5"].value == "ALLEMAGNE"
    assert onglet["G5"].value == "42000"
    assert onglet["H5"].value == "commentaire F2"
    assert onglet["I4"].value == "Société à responsabilité limitée (SARL)"
    assert onglet["J4"].value == "OUI"
    assert onglet["K4"].value == "Culture du riz"
    assert onglet["K5"].value == "Culture de plantes à fibres"
    assert onglet["L4"].value == 54321
    assert onglet["M4"].value == 98765
    assert onglet["N4"].value == "Équivalents temps plein (ETP)"
    assert onglet["O4"].value == 42.0
    assert onglet["P4"].value == "ALLEMAGNE"
    assert onglet["P5"].value == "FRANCE"
    assert onglet["Q4"].value == 1
    assert onglet["R4"].value == "site 1"
    assert onglet["S4"].value == "adresse site 1"
    assert onglet["T4"].value == "75000"
    assert onglet["U4"].value == "Paris"
    assert onglet["V4"].value == "FRANCE"
    assert onglet["W4"].value == "[2.294381,48.858099]"
    assert onglet["Q5"].value == 2
    assert onglet["R5"].value == "site 2"
    assert onglet["S5"].value == "adresse site 2"
    assert onglet["T5"].value == "12345"
    assert onglet["U5"].value == "Berlin"
    assert onglet["V5"].value == "ALLEMAGNE"
    assert onglet["W5"].value == "[3,3]"
    assert onglet["X4"].value == "label"
    assert onglet["Y4"].value == "émetteur 1"
    assert onglet["Z4"].value == "2000-12-12"
    assert onglet["AA4"].value == "10/20"
    assert onglet["AB4"].value == "un commentaire"
    assert onglet["X5"].value == "certif"
    assert onglet["Y5"].value == "émetteur 2"
    assert onglet["Z5"].value == "2020-04-05"
    assert onglet["AA5"].value == "bon"
    assert not onglet["AB5"].value


def test_telechargement_d_un_rapport_vsme_au_format_xlsx_B1_avec_indicateurs_non_pertinents(
    client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    rapport_vsme = RapportVSME.objects.create(entreprise=entreprise, annee=2025)
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-24-b",
        data={"non_pertinent": True, "omission_informations": []},
    )
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-25",
        data={"non_pertinent": True, "certifications": []},
    )
    client.force_login(alice)

    response = client.get(f"/vsme/{rapport_vsme.id}/export/xlsx")

    assert response["Content-Disposition"] == "filename=vsme.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook["B1"]
    assert not onglet["B4"].value
    assert not onglet["T4"].value
    assert not onglet["U4"].value
    assert not onglet["V4"].value
    assert not onglet["W4"].value
    assert not onglet["X4"].value


def test_telechargement_d_un_rapport_vsme_au_format_xlsx_B2(
    client, rapport_vsme, alice
):
    rapport_vsme.indicateurs.create(
        rapport_vsme=rapport_vsme,
        schema_id="B2-26-p1",
        data={"non_pertinent": False, "participation_gouvernance": "PARTICIPATION"},
    )
    rapport_vsme.indicateurs.create(
        rapport_vsme=rapport_vsme,
        schema_id="B2-26-p2",
        data={"non_pertinent": False, "investissement_economie_sociale": 222},
    )
    rapport_vsme.indicateurs.create(
        rapport_vsme=rapport_vsme,
        schema_id="B2-26-p3",
        data={"non_pertinent": False, "limites_distribution_profits": "LIMITES"},
    )
    # Active les indicateurs spécifiques aux coopératives B2-26-p1 B2-26-p2 et B2-26-p3
    rapport_vsme.indicateurs.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-24-e-i",
        data={"forme_juridique": "54", "coopérative": True},
    )

    rapport_vsme.indicateurs.create(
        rapport_vsme=rapport_vsme,
        schema_id="B2-26",
        data={
            "declaration_durabilite": {
                "changement_climatique": {
                    "pratiques": True,
                    "accessibles": True,
                    "cibles": False,
                },
                "pollution": {
                    "pratiques": False,
                    "accessibles": False,
                    "cibles": False,
                },
                "eau": {"pratiques": False, "accessibles": False, "cibles": True},
                "biodiversite": {
                    "pratiques": False,
                    "accessibles": False,
                    "cibles": False,
                },
                "economie_circulaire": {
                    "pratiques": False,
                    "accessibles": False,
                    "cibles": False,
                },
                "personnel": {
                    "pratiques": False,
                    "accessibles": False,
                    "cibles": False,
                },
                "travailleurs": {
                    "pratiques": False,
                    "accessibles": False,
                    "cibles": False,
                },
                "communautes": {
                    "pratiques": False,
                    "accessibles": False,
                    "cibles": False,
                },
                "consommateurs": {
                    "pratiques": False,
                    "accessibles": False,
                    "cibles": False,
                },
                "conduite_affaires": {
                    "pratiques": False,
                    "accessibles": True,
                    "cibles": True,
                },
            }
        },
    )
    client.force_login(alice)

    response = client.get(f"/vsme/{rapport_vsme.id}/export/xlsx")

    assert response["Content-Disposition"] == "filename=vsme.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet_b2 = workbook["B2"]
    assert onglet_b2["A4"].value == "PARTICIPATION"
    assert onglet_b2["B4"].value == 222
    assert onglet_b2["C4"].value == "LIMITES"
    assert onglet_b2["E5"].value == "OUI"
    assert onglet_b2["F5"].value == "OUI"
    assert onglet_b2["G5"].value == "NON"
    assert onglet_b2["E14"].value == "NON"
    assert onglet_b2["F14"].value == "OUI"
    assert onglet_b2["G14"].value == "OUI"


def test_telechargement_d_un_rapport_vsme_au_format_xlsx_B2_avec_indicateurs_non_applicables(
    client, rapport_vsme, alice
):
    rapport_vsme.indicateurs.create(
        rapport_vsme=rapport_vsme,
        schema_id="B2-26-p1",
        data={"non_pertinent": False, "participation_gouvernance": "PARTICIPATION"},
    )
    rapport_vsme.indicateurs.create(
        rapport_vsme=rapport_vsme,
        schema_id="B2-26-p2",
        data={"non_pertinent": False, "investissement_economie_sociale": 222},
    )
    rapport_vsme.indicateurs.create(
        rapport_vsme=rapport_vsme,
        schema_id="B2-26-p3",
        data={"non_pertinent": False, "limites_distribution_profits": "LIMITES"},
    )
    # Désactive les indicateurs spécifiques aux coopératives B2-26-p1 B2-26-p2 et B2-26-p3
    rapport_vsme.indicateurs.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-24-e-i",
        data={"forme_juridique": "54", "coopérative": False},
    )

    rapport_vsme.indicateurs.create(
        schema_id="B2-26",
        data={
            "declaration_durabilite": {
                "changement_climatique": {
                    "pratiques": True,
                    "accessibles": True,
                    "cibles": False,
                },
                "pollution": {
                    "pratiques": False,
                    "accessibles": False,
                    "cibles": False,
                },
                "eau": {"pratiques": False, "accessibles": False, "cibles": True},
                "biodiversite": {
                    "pratiques": False,
                    "accessibles": False,
                    "cibles": False,
                },
                "economie_circulaire": {
                    "pratiques": False,
                    "accessibles": False,
                    "cibles": False,
                },
                "personnel": {
                    "pratiques": False,
                    "accessibles": False,
                    "cibles": False,
                },
                "travailleurs": {
                    "pratiques": False,
                    "accessibles": False,
                    "cibles": False,
                },
                "communautes": {
                    "pratiques": False,
                    "accessibles": False,
                    "cibles": False,
                },
                "consommateurs": {
                    "pratiques": False,
                    "accessibles": False,
                    "cibles": False,
                },
                "conduite_affaires": {
                    "pratiques": False,
                    "accessibles": True,
                    "cibles": True,
                },
            }
        },
    )
    client.force_login(alice)

    response = client.get(f"/vsme/{rapport_vsme.id}/export/xlsx")

    workbook = load_workbook(filename=BytesIO(response.content))
    onglet_b2 = workbook["B2"]
    assert not onglet_b2["A4"].value
    assert not onglet_b2["B4"].value
    assert not onglet_b2["C4"].value
    assert onglet_b2["E5"].value == "OUI"
    assert onglet_b2["F5"].value == "OUI"
    assert onglet_b2["G5"].value == "NON"
    assert onglet_b2["E14"].value == "NON"
    assert onglet_b2["F14"].value == "OUI"
    assert onglet_b2["G14"].value == "OUI"


def test_telechargement_d_un_rapport_vsme_au_format_xlsx_B4(
    client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    rapport_vsme = RapportVSME.objects.create(entreprise=entreprise, annee=2025)
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B4-32-p1",
        data={
            "non_pertinent": False,
            "pollution_air": [
                {"polluant": "Anthracène", "unite": "kilos", "valeur": 11.0},
                {"polluant": "Autre", "unite": "tonnes", "valeur": 22.0},
            ],
        },
    )
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B4-32-p2",
        data={
            "non_pertinent": False,
            "pollution_eau": [
                {"polluant": "Chlorfenvinphos", "unite": "tonnes", "valeur": 33.0},
                {"polluant": "Chlorure de vinyle", "unite": "kilos", "valeur": 44.0},
            ],
        },
    )
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B4-32-p3",
        data={
            "non_pertinent": False,
            "pollution_sols": [
                {"polluant": "Alachlore", "unite": "kilos", "valeur": 55.0},
                {"polluant": "divers", "unite": "tonnes", "valeur": 66.0},
            ],
        },
    )
    client.force_login(alice)

    response = client.get(f"/vsme/{rapport_vsme.id}/export/xlsx")

    assert response["Content-Disposition"] == "filename=vsme.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook["B4"]
    assert onglet["A4"].value == "Anthracène"
    assert onglet["B4"].value == "kg"
    assert onglet["C4"].value == 11
    assert onglet["A5"].value == "Autre"
    assert onglet["B5"].value == "tonnes"
    assert onglet["C5"].value == 22
    assert onglet["D4"].value == "Chlorfenvinphos"
    assert onglet["E4"].value == "tonnes"
    assert onglet["F4"].value == 33
    assert onglet["D5"].value == "Chlorure de vinyle"
    assert onglet["E5"].value == "kg"
    assert onglet["F5"].value == 44
    assert onglet["G4"].value == "Alachlore"
    assert onglet["H4"].value == "kg"
    assert onglet["I4"].value == 55
    assert onglet["G5"].value == "divers"
    assert onglet["H5"].value == "tonnes"
    assert onglet["I5"].value == 66


def test_telechargement_d_un_rapport_vsme_au_format_xlsx_B5(
    client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    rapport_vsme = RapportVSME.objects.create(entreprise=entreprise, annee=2025)
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-24-e-vii",
        data={
            "sites": [
                {
                    "id_site": 1,
                    "nom_site": "Usine",
                    "adresse": "adresse",
                    "code_postal": "33000",
                    "ville": "Bordeaux",
                    "pays": "FRA",
                    "geolocalisation": "[2.294381,48.858099]",
                }
            ]
        },
    )
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B5-33",
        data={
            "non_pertinent": False,
            "unite_superficie": "m2",
            "sites_zones_sensibles": [
                {
                    "id_site": "1",
                    "superficie": 999.0,
                    "dans_zone_sensible": True,
                    "proximite_zone_sensible": True,
                }
            ],
        },
    )
    client.force_login(alice)

    response = client.get(f"/vsme/{rapport_vsme.id}/export/xlsx")

    assert response["Content-Disposition"] == "filename=vsme.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook["B5"]
    assert onglet["A4"].value == "mètres carrés (m²)"
    assert onglet["B4"].value == "1"
    assert onglet["C4"].value == 999.0
    assert onglet["D4"].value == "OUI"
    assert onglet["E4"].value == "OUI"


def test_telechargement_d_un_rapport_vsme_au_format_xlsx_B6(
    client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    rapport_vsme = RapportVSME.objects.create(entreprise=entreprise, annee=2025)
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B6-35",
        data={
            "total_prelevements_eau": 2222,
            "total_prelevements_eau_sites_sensibles": 1111,
        },
    )
    client.force_login(alice)

    response = client.get(f"/vsme/{rapport_vsme.id}/export/xlsx")

    assert response["Content-Disposition"] == "filename=vsme.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook["B6"]
    assert onglet["A4"].value == 2222
    assert onglet["B4"].value == 1111


def test_telechargement_d_un_rapport_vsme_au_format_xlsx_B7(
    client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    rapport_vsme = RapportVSME.objects.create(entreprise=entreprise, annee=2025)
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B7-37",
        data={"non_pertinent": False, "economie_circulaire": "PRINCIPES"},
    )
    client.force_login(alice)

    response = client.get(f"/vsme/{rapport_vsme.id}/export/xlsx")

    assert response["Content-Disposition"] == "filename=vsme.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook["B7"]
    assert onglet["A4"].value == "PRINCIPES"


def test_telechargement_d_un_rapport_vsme_au_format_xlsx_B8(
    client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    rapport_vsme = RapportVSME.objects.create(entreprise=entreprise, annee=2025)
    # Indicateur pays
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B1-24-e-vi",
        data={"pays": ["FIN", "FRA"]},
    )
    # Indicateur effectifs par pays
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B8-39-c",
        data={
            "effectifs_pays": {
                "FIN": {"nombre_salaries": 30},
                "FRA": {"nombre_salaries": 12},
            }
        },
    )
    # Indicateur effectifs par type de contrat
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B8-39-a",
        data={
            "effectifs_type_de_contrat": {
                "contrat_permanent": {"nombre_salaries": 40.5},
                "contrat_temporaire": {"nombre_salaries": 1.5},
            }
        },
    )
    client.force_login(alice)

    response = client.get(f"/vsme/{rapport_vsme.id}/export/xlsx")

    assert response["Content-Disposition"] == "filename=vsme.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook["B8"]
    assert onglet["A4"].value == 40.5
    assert onglet["B4"].value == 1.5
    assert onglet["G4"].value == "FINLANDE"
    assert onglet["H4"].value == 30
    assert onglet["G5"].value == "FRANCE"
    assert onglet["H5"].value == 12


def test_telechargement_d_un_rapport_vsme_au_format_xlsx_B9(
    client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    rapport_vsme = RapportVSME.objects.create(entreprise=entreprise, annee=2025)
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B9-41a",
        data={"nombre_accidents_travail": 55, "taux_accidents_travail": 12.3},
    )
    client.force_login(alice)

    response = client.get(f"/vsme/{rapport_vsme.id}/export/xlsx")

    assert response["Content-Disposition"] == "filename=vsme.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook["B9"]
    assert onglet["A4"].value == 55
    assert onglet["B4"].value == 12.3


def test_telechargement_d_un_rapport_vsme_au_format_xlsx_B10(
    client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    rapport_vsme = RapportVSME.objects.create(entreprise=entreprise, annee=2025)
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B10-42-a",
        data={
            "respect_salaire_minimum": True,
        },
    )
    client.force_login(alice)

    response = client.get(f"/vsme/{rapport_vsme.id}/export/xlsx")

    assert response["Content-Disposition"] == "filename=vsme.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook["B10"]
    assert onglet["A4"].value == "OUI"


def test_telechargement_d_un_rapport_vsme_au_format_xlsx_B10_avec_champ_calculé(
    client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    rapport_vsme = RapportVSME.objects.create(entreprise=entreprise, annee=2025)
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B10-42-b",
        data={
            "remuneration_horaire_hommes": 2,
            "remuneration_horaire_femmes": 1,
        },
    )
    client.force_login(alice)

    response = client.get(f"/vsme/{rapport_vsme.id}/export/xlsx")

    assert response["Content-Disposition"] == "filename=vsme.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook["B10"]
    assert onglet["D4"].value == 50  # résultat de ecart_remuneration_hommes_femmes


def test_telechargement_d_un_rapport_vsme_au_format_xlsx_B11(
    client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    rapport_vsme = RapportVSME.objects.create(entreprise=entreprise, annee=2025)
    Indicateur.objects.create(
        rapport_vsme=rapport_vsme,
        schema_id="B11-43-p1",
        data={"non_pertinent": False, "nombre_condamnations": 0},
    )
    client.force_login(alice)

    response = client.get(f"/vsme/{rapport_vsme.id}/export/xlsx")

    assert response["Content-Disposition"] == "filename=vsme.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook["B11"]
    assert onglet["A4"].value == 0


def test_telechargement_d_un_rapport_vsme_inexistant(client, entreprise_factory, alice):
    entreprise = entreprise_factory(utilisateur=alice)
    client.force_login(alice)

    response = client.get(f"/vsme/42/export/xlsx")

    assert response.status_code == 404


def test_telechargement_d_un_rapport_vsme_redirige_vers_la_connexion_si_non_connecté(
    client, entreprise_factory, alice
):
    entreprise = entreprise_factory(utilisateur=alice)
    rapport_vsme = RapportVSME.objects.create(entreprise=entreprise, annee=2025)

    response = client.get(f"/vsme/{rapport_vsme.id}/export/xlsx")

    assert response.status_code == 302

    response = client.get(
        reverse("analyseia:synthese_resultat", args=[entreprise.siren, 42]),
    )

    assert response.status_code == 302

from datetime import date
from datetime import datetime
from io import BytesIO

import pytest
from django.contrib.messages import WARNING
from django.urls import reverse
from openpyxl import load_workbook
from pytest_django.asserts import assertTemplateUsed

from habilitations.models import attach_user_to_entreprise
from reglementations.enums import EtapeCSRD
from reglementations.models.csrd import Enjeu
from reglementations.models.csrd import RapportCSRD


def test_l_espace_csrd_n_est_pas_public(client, alice, entreprise_factory):
    entreprise = entreprise_factory()

    url = f"/csrd/guide/{entreprise.siren}"
    response = client.get(url)

    assert response.status_code == 302
    connexion_url = reverse("users:login")
    assert response.url == f"{connexion_url}?next={url}"

    client.force_login(alice)
    response = client.get(url)

    assert response.status_code == 403

    attach_user_to_entreprise(alice, entreprise, "Présidente")
    response = client.get(url)

    assert response.status_code == 200
    assert response.templates[0].name == "reglementations/espace_csrd/index.html"


def test_espace_csrd_sans_siren_redirige_vers_celui_de_l_entreprise_courante(
    client, alice, entreprise_factory
):
    entreprise = entreprise_factory()
    attach_user_to_entreprise(alice, entreprise, "Présidente")
    client.force_login(alice)

    url = "/csrd"
    response = client.get(url)

    assert response.status_code == 302
    assert response.url == f"/csrd/guide/{entreprise.siren}"


def test_espace_csrd_sans_siren_et_sans_entreprise(client, alice):
    # Cas limite où un utilisateur n'est rattaché à aucune entreprise
    client.force_login(alice)

    url = "/csrd"
    response = client.get(url, follow=True)

    assert response.status_code == 200
    assert response.redirect_chain == [(reverse("entreprises:entreprises"), 302)]
    messages = list(response.context["messages"])
    assert messages[0].level == WARNING
    assert (
        messages[0].message
        == "Commencez par ajouter une entreprise à votre compte utilisateur avant d'accéder à l'espace Rapport de Durabilité"
    )


def test_espace_csrd_avec_siren_inexistant(client, alice):
    client.force_login(alice)

    url = "/csrd/yolo"
    response = client.get(url)

    assert response.status_code == 404


@pytest.mark.parametrize(
    "etape",
    [
        "/csrd/guide/{siren}/phase-1",
        "/csrd/guide/{siren}/phase-1/etape-1",
        "/csrd/guide/{siren}/phase-1/etape-1-1",
        "/csrd/guide/{siren}/phase-1/etape-1-1",
        "/csrd/guide/{siren}/phase-1/etape-1-3",
        "/csrd/guide/{siren}/phase-1/etape-2",
        "/csrd/guide/{siren}/phase-1/etape-2-1",
        "/csrd/guide/{siren}/phase-1/etape-2-2",
        "/csrd/guide/{siren}/phase-1/etape-3",
        "/csrd/guide/{siren}/phase-1/etape-3-1",
        "/csrd/guide/{siren}/phase-1/etape-3-2",
        "/csrd/guide/{siren}/phase-1/etape-3-3",
        "/csrd/guide/{siren}/phase-2",
        "/csrd/guide/{siren}/phase-2/etape-1",
        "/csrd/guide/{siren}/phase-3",
        "/csrd/guide/{siren}/phase-3/etape-1",
    ],
)
def test_guide_de_la_csrd_par_etape(etape, client, alice, entreprise_factory):
    entreprise = entreprise_factory()
    url = etape.format(siren=entreprise.siren)

    response = client.get(url)

    assert response.status_code == 302
    connexion_url = reverse("users:login")
    assert response.url == f"{connexion_url}?next={url}"

    attach_user_to_entreprise(alice, entreprise, "Présidente")
    client.force_login(alice)
    response = client.get(url)

    assert response.status_code == 200
    context = response.context
    assert context["entreprise"] == entreprise
    assert "phase" in context
    assert "etape" in context
    assert "sous_etape" in context

    etape_inexistante = f"/csrd/guide/{entreprise.siren}/phase-4/etape-1"
    response = client.get(etape_inexistante)

    assert response.status_code == 404


@pytest.mark.parametrize(
    "etape",
    [
        "/csrd/{siren}/etape-introduction",
        "/csrd/{siren}/etape-selection-enjeux",
        "/csrd/{siren}/etape-analyse-materialite",
        "/csrd/{siren}/etape-collection-donnees-entreprise",
        "/csrd/{siren}/etape-redaction-rapport-durabilite",
    ],
)
def test_gestion_de_la_csrd(etape, client, alice, entreprise_factory):
    entreprise = entreprise_factory()
    url = etape.format(siren=entreprise.siren)

    response = client.get(url)

    assert response.status_code == 302
    connexion_url = reverse("users:login")
    assert response.url == f"{connexion_url}?next={url}"

    attach_user_to_entreprise(alice, entreprise, "Présidente")
    client.force_login(alice)
    response = client.get(url)

    assert response.status_code == 200
    context = response.context
    assert context["entreprise"] == entreprise
    if etape.endswith("introduction"):
        assertTemplateUsed(response, "reglementations/csrd/etape-introduction.html")
    elif etape.endswith("selection-enjeux"):
        assertTemplateUsed(response, "reglementations/csrd/etape-selection-enjeux.html")
    elif etape.endswith("analyse-materialite"):
        assertTemplateUsed(
            response, "reglementations/csrd/etape-analyse-materialite.html"
        )
    elif etape.endswith("collection-donnees-entreprise"):
        assertTemplateUsed(
            response, "reglementations/csrd/etape-collection-donnees-entreprise.html"
        )
    elif etape.endswith("redaction-rapport-durabilite"):
        assertTemplateUsed(
            response, "reglementations/csrd/etape-redaction-rapport-durabilite.html"
        )

    rapport_csrd = RapportCSRD.objects.get(proprietaire=alice, entreprise=entreprise)
    NOMBRE_ENJEUX = 103
    assert len(rapport_csrd.enjeux.all()) == NOMBRE_ENJEUX


def test_étape_inexistante_de_la_csrd(client, alice, entreprise_factory):
    entreprise = entreprise_factory()
    attach_user_to_entreprise(alice, entreprise, "Présidente")
    client.force_login(alice)
    etape_inexistante = f"/csrd/{entreprise.siren}/etape-4"

    response = client.get(etape_inexistante)

    assert response.status_code == 404


@pytest.mark.parametrize(
    "etape",
    [
        "introduction",
        "selection-enjeux",
        "analyse-materialite",
        "collection-donnees-entreprise",
    ],
)
def test_enregistrement_de_l_étape_de_la_csrd(etape, client, alice, entreprise_factory):
    entreprise = entreprise_factory()
    habilitation = attach_user_to_entreprise(alice, entreprise, "Présidente")
    RapportCSRD.objects.create(
        entreprise=entreprise,
        proprietaire=alice,
        annee=date.today().year,
    )
    client.force_login(alice)
    url = "/csrd/{siren}/etape-{etape}".format(
        siren=entreprise.siren, etape=EtapeCSRD.id_suivant(etape)
    )

    response = client.post(url, follow=True)

    rapport_csrd = RapportCSRD.objects.get(proprietaire=alice, entreprise=entreprise)
    assert rapport_csrd.etape_validee == etape


@pytest.mark.parametrize(
    "etape",
    [
        "introduction",
        "selection-enjeux",
        "analyse-materialite",
        "collection-donnees-entreprise",
    ],
)
def test_enregistrement_de_l_étape_de_la_csrd_retourne_une_404_si_aucune_CSRD(
    etape, client, alice, entreprise_factory
):
    entreprise = entreprise_factory()
    attach_user_to_entreprise(alice, entreprise, "Présidente")
    client.force_login(alice)
    url = "/csrd/{siren}/etape-{etape}".format(
        siren=entreprise.siren, etape=EtapeCSRD.id_suivant(etape)
    )

    response = client.post(url, follow=True)

    assert response.status_code == 404
    assert (
        RapportCSRD.objects.filter(proprietaire=alice, entreprise=entreprise).count()
        == 0
    )


def test_visualisation_des_enjeux(client, alice, entreprise_non_qualifiee):
    attach_user_to_entreprise(alice, entreprise_non_qualifiee, "Présidente")
    csrd = RapportCSRD.objects.create(
        proprietaire=alice,
        entreprise=entreprise_non_qualifiee,
        annee=f"{datetime.now():%Y}",
    )
    client.force_login(alice)

    response = client.get(f"/csrd/fragments/selection_enjeux/{csrd.id}/ESRS_E1", {})

    assert response.status_code == 200
    assert "<!-- fragment enjeux -->" in response.content.decode("utf-8")


def test_selection_et_deselection_d_enjeux(client, alice, entreprise_non_qualifiee):
    # update : les enjeux sont désormais sélectionnés par défaut
    attach_user_to_entreprise(alice, entreprise_non_qualifiee, "Présidente")
    csrd = RapportCSRD.objects.create(
        proprietaire=alice,
        entreprise=entreprise_non_qualifiee,
        annee=f"{datetime.now():%Y}",
    )
    enjeux = csrd.enjeux.all()
    enjeu_adaptation = enjeux[0]
    enjeu_attenuation = enjeux[1]
    enjeu_energie = enjeux[2]
    client.force_login(alice)

    response = client.post(
        f"/csrd/fragments/selection_enjeux/{csrd.id}/ESRS_E1",
        {"enjeux": [enjeu_adaptation.id, enjeu_energie.id]},
    )

    enjeu_adaptation.refresh_from_db()
    enjeu_attenuation.refresh_from_db()
    enjeu_energie.refresh_from_db()

    assert [enjeu_attenuation] == list(Enjeu.objects.filter(selection=False))
    assert response.status_code == 200

    context = response.context

    assert context["csrd"] == csrd
    assert "<!-- fragment esrs -->" in response.content.decode("utf-8")


def test_deselection_d_un_enjeu(client, alice, entreprise_non_qualifiee):
    attach_user_to_entreprise(alice, entreprise_non_qualifiee, "Présidente")
    csrd = RapportCSRD.objects.create(
        proprietaire=alice,
        entreprise=entreprise_non_qualifiee,
        annee=f"{datetime.now():%Y}",
    )
    enjeux = csrd.enjeux.all()
    enjeu = enjeux[0]
    enjeu.selection = True
    enjeu.save()
    client.force_login(alice)

    response = client.post(
        f"/csrd/fragments/deselection_enjeu/{enjeu.id}",
    )

    enjeu.refresh_from_db()
    assert not enjeu.selection
    assert response.status_code == 200


def test_liste_des_enjeux_csrd_au_format_xlsx(client, alice, entreprise_non_qualifiee):
    attach_user_to_entreprise(alice, entreprise_non_qualifiee, "Présidente")
    csrd = RapportCSRD.objects.create(
        proprietaire=alice,
        entreprise=entreprise_non_qualifiee,
        annee=f"{datetime.now():%Y}",
    )
    enjeux = csrd.enjeux.all()
    enjeu_adaptation = enjeux[0]
    enjeu_attenuation = enjeux[1]
    enjeu_attenuation.selection = True
    enjeu_attenuation.save()
    enjeu_energie = enjeux[2]
    client.force_login(alice)

    response = client.get(
        f"/csrd/{entreprise_non_qualifiee.siren}/enjeux.xlsx",
    )

    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )


def test_liste_des_enjeux_csrd__au_format_xlsx_retourne_une_404_si_entreprise_inexistante(
    client, alice
):
    client.force_login(alice)

    response = client.get(
        f"/csrd/000000001/enjeux.xlsx",
    )

    assert response.status_code == 404


def test_liste_des_enjeux_csrd_au_format_xlsx_retourne_une_404_si_habilitation_inexistante(
    client, alice, entreprise_non_qualifiee
):
    client.force_login(alice)

    response = client.get(
        f"/csrd/{entreprise_non_qualifiee.siren}/enjeux.xlsx",
    )

    assert response.status_code == 404


def test_liste_des_enjeux_csrd_au_format_xlsx_retourne_une_404_si_csrd_inexistante(
    client, alice, entreprise_non_qualifiee
):
    attach_user_to_entreprise(alice, entreprise_non_qualifiee, "Présidente")
    client.force_login(alice)

    response = client.get(
        f"/csrd/{entreprise_non_qualifiee.siren}/enjeux.xlsx",
    )

    assert response.status_code == 404


def test_liste_des_enjeux_csrd(client, alice, entreprise_non_qualifiee):
    attach_user_to_entreprise(alice, entreprise_non_qualifiee, "Présidente")
    csrd = RapportCSRD.objects.create(
        proprietaire=alice,
        entreprise=entreprise_non_qualifiee,
        annee=f"{datetime.now():%Y}",
    )
    enjeux = csrd.enjeux.all()

    enjeu_attenuation = enjeux[1]
    enjeu_attenuation.selection = False
    enjeu_attenuation.save()

    assert (
        enjeux.filter(selection=False).count() == 1
    ), "Un des enjeux doit être désélectionné"

    client.force_login(alice)
    response = client.get(f"/csrd/fragments/liste_enjeux_selectionnes/{csrd.id}/1")

    assert response.status_code == 200

    context = response.context

    assert (
        len(
            context["enjeux_par_esg"]["environnement"][
                "ESRS E1 - Changement climatique"
            ]
        )
        == 2
    ), "Un des enjeux doit être désélectionné"
    assert "<!-- fragment liste des enjeux sélectionnés -->" in response.content.decode(
        "utf-8"
    )


def test_datapoints_pour_enjeux_materiels_au_format_xlsx(
    client, alice, entreprise_non_qualifiee
):
    attach_user_to_entreprise(alice, entreprise_non_qualifiee, "Présidente")
    csrd = RapportCSRD.objects.create(
        proprietaire=alice,
        entreprise=entreprise_non_qualifiee,
        annee=f"{datetime.now():%Y}",
    )
    enjeux = csrd.enjeux.all()
    enjeu_attenuation = enjeux[1]
    enjeu_attenuation.selection = True
    enjeu_attenuation.materiel = True
    enjeu_attenuation.save()
    esrs_materielle = enjeu_attenuation.esrs
    client.force_login(alice)

    response = client.get(
        f"/csrd/{entreprise_non_qualifiee.siren}/datapoints.xlsx",
    )

    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    noms_onglet = workbook.get_sheet_names()
    assert esrs_materielle.replace("_", " ") in noms_onglet
    assert "Index" in noms_onglet
    assert "ESRS 2" in noms_onglet
    assert "ESRS2 MDR" in noms_onglet
    assert "ESRS G1" not in noms_onglet


def test_datapoints_pour_enjeux_non_materiels_au_format_xlsx(
    client, alice, entreprise_non_qualifiee
):
    attach_user_to_entreprise(alice, entreprise_non_qualifiee, "Présidente")
    csrd = RapportCSRD.objects.create(
        proprietaire=alice,
        entreprise=entreprise_non_qualifiee,
        annee=f"{datetime.now():%Y}",
    )
    enjeux = csrd.enjeux.all()

    # enjeu de l'ESRS_E1:
    enjeu_attenuation = enjeux[1]
    enjeu_attenuation.selection = True
    enjeu_attenuation.materiel = True
    enjeu_attenuation.save()
    esrs_materielle = enjeu_attenuation.esrs
    client.force_login(alice)

    # note : les enjeux affichés dans le fichier "non-matériels"
    # doivent au préalable avoir été sélectionnés

    # enjeu de l'ESRS_G1:
    enjeux_G1 = enjeux.filter(esrs="ESRS_G1").first()
    enjeux_G1.selection = True
    enjeux_G1.materiel = False  # et pas None
    enjeux_G1.save()

    response = client.get(
        f"/csrd/{entreprise_non_qualifiee.siren}/datapoints.xlsx?materiel=false",
    )

    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    noms_onglet = workbook.sheetnames

    assert esrs_materielle.replace("_", " ") not in noms_onglet
    assert "Index" in noms_onglet
    assert "ESRS 2" in noms_onglet
    assert "ESRS2 MDR" in noms_onglet
    assert "ESRS G1" in noms_onglet


def test_datapoints_csrd__au_format_xlsx_retourne_une_404_si_entreprise_inexistante(
    client, alice
):
    client.force_login(alice)

    response = client.get(
        f"/csrd/000000001/datapoints.xlsx",
    )

    assert response.status_code == 404


def test_datapoints_csrd_au_format_xlsx_retourne_une_404_si_habilitation_inexistante(
    client, alice, entreprise_non_qualifiee
):
    client.force_login(alice)

    response = client.get(
        f"/csrd/{entreprise_non_qualifiee.siren}/datapoints.xlsx",
    )

    assert response.status_code == 404


def test_datapoints_csrd_au_format_xlsx_retourne_une_404_si_csrd_inexistante(
    client, alice, entreprise_non_qualifiee
):
    attach_user_to_entreprise(alice, entreprise_non_qualifiee, "Présidente")
    client.force_login(alice)

    response = client.get(
        f"/csrd/{entreprise_non_qualifiee.siren}/datapoints.xlsx",
    )

    assert response.status_code == 404

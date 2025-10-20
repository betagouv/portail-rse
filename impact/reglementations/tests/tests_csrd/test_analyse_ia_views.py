from io import BytesIO

from django.conf import settings
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import load_workbook

from analyseia.models import AnalyseIA
from api.exceptions import APIError

CONTENU_PDF = b"%PDF-1.4\n%\xd3\xeb\xe9\xe1\n1 0 obj\n<</Title (CharteEngagements"


# note : les rapports CSRD sont uniquement officiels désormais
# Alice est l'utilisateur propriétaire rattaché aux entreprises de test


def test_ajout_document_par_utilisateur_autorise(client, csrd, alice):
    client.force_login(alice)
    fichier = SimpleUploadedFile("test.pdf", CONTENU_PDF)

    response = client.post(
        f"/csrd/{csrd.id}/ajout_document",
        {"fichier": fichier},
    )

    assert response.status_code == 302
    assert response.url == reverse(
        "reglementations:gestion_csrd",
        kwargs={
            "siren": csrd.entreprise.siren,
            "id_etape": "analyse-ecart",
        },
    )
    assert csrd.analyses_ia.count() == 1
    assert csrd.analyses_ia.first().nom == "test.pdf"


def test_ajout_document_par_utilisateur_non_autorise(client, csrd, bob):
    client.force_login(bob)
    fichier = SimpleUploadedFile("test.pdf", CONTENU_PDF)

    response = client.post(
        f"/csrd/{csrd.id}/ajout_document",
        {"fichier": fichier},
    )

    assert response.status_code == 403
    assert csrd.analyses_ia.count() == 0


def test_ajout_document_sur_csrd_inexistante(client, alice):
    client.force_login(alice)
    fichier = SimpleUploadedFile("test.pdf", CONTENU_PDF)

    response = client.post(
        "/csrd/42/ajout_document",
        {"fichier": fichier},
    )

    assert response.status_code == 404
    assert AnalyseIA.objects.count() == 0


def test_ajout_document_sans_extension_pdf(client, csrd, alice):
    client.force_login(alice)
    fichier = SimpleUploadedFile("test.odt", b"libre office writer data")

    response = client.post(
        f"/csrd/{csrd.id}/ajout_document",
        {"fichier": fichier},
    )

    assert response.status_code == 400
    assert csrd.analyses_ia.count() == 0


def test_ajout_document_dont_le_contenu_n_est_pas_du_pdf(client, csrd, alice):
    client.force_login(alice)
    fichier = SimpleUploadedFile("test.pdf", b"pas un pdf")

    response = client.post(
        f"/csrd/{csrd.id}/ajout_document",
        {"fichier": fichier},
    )

    assert response.status_code == 400
    assert csrd.analyses_ia.count() == 0


def test_suppression_document_par_utilisateur_autorise(client, document, alice):
    client.force_login(alice)
    rapport_csrd = document.rapports_csrd.select_related("entreprise").first()

    url = reverse(
        "reglementations:suppression_document", kwargs={"id_document": document.id}
    )
    response = client.post(url)

    assert response.status_code == 302
    assert response.url == reverse(
        "reglementations:gestion_csrd",
        kwargs={
            "siren": rapport_csrd.entreprise.siren,
            "id_etape": "analyse-ecart",
        },
    )
    assert AnalyseIA.objects.count() == 0


def test_suppression_document_par_utilisateur_non_autorise(client, document, bob):
    client.force_login(bob)

    url = reverse(
        "reglementations:suppression_document", kwargs={"id_document": document.id}
    )
    response = client.post(url)

    assert response.status_code == 403
    assert AnalyseIA.objects.count() == 1


def test_suppression_document_inexistant(client, document, alice):
    client.force_login(alice)

    url = reverse("reglementations:suppression_document", kwargs={"id_document": 42})
    response = client.post(url)

    assert response.status_code == 404
    assert AnalyseIA.objects.count() == 1


def test_lancement_d_analyse_IA(client, mock_api_analyse_ia, document, alice):
    client.force_login(alice)
    rapport_csrd = document.rapports_csrd.select_related("entreprise").first()

    response = client.post(
        f"/ESRS-predict/{document.id}/{rapport_csrd.id}/start",
        follow=True,
    )

    callback_url = response.wsgi_request.build_absolute_uri(
        reverse(
            "reglementations:etat_analyse_IA",
            kwargs={
                "id_document": document.id,
                "csrd_id": rapport_csrd.id,
            },
        )
    )

    mock_api_analyse_ia.assert_called_once_with(
        document.id, document.fichier.url, callback_url
    )
    document.refresh_from_db()
    assert document.etat == "pending"
    assert response.status_code == 200
    assert "L'analyse a bien été lancée." in response.content.decode("utf-8")


def test_lancement_d_anlyse_IA_redirige_vers_la_connexion_si_non_connecté(
    client, mock_api_analyse_ia, document
):
    response = client.post(
        f"/ESRS-predict/{document.id}/start",
    )

    assert response.status_code == 302
    assert not document.etat
    assert not mock_api_analyse_ia.called


def test_lancement_d_analyse_IA_erreur_API(
    client, mock_api_analyse_ia, document, alice
):
    rapport_csrd = document.rapports_csrd.select_related("entreprise").first()
    message_erreur = (
        "Le service est actuellement indisponible. Merci de réessayer plus tard."
    )
    mock_api_analyse_ia.side_effect = APIError(message_erreur)
    client.force_login(alice)

    response = client.post(
        f"/ESRS-predict/{document.id}/{rapport_csrd.id}/start", follow=True
    )

    document.refresh_from_db()
    assert not document.etat
    content = response.content.decode("utf-8")
    assert message_erreur in content


def test_serveur_IA_envoie_l_etat_d_avancement_de_l_analyse_1(
    client, document, mailoutbox
):
    rapport_csrd = document.rapports_csrd.select_related("entreprise").first()

    client.post(
        f"/ESRS-predict/{document.id}/{rapport_csrd.id}",
        {
            "status": "processing",
        },
    )

    document.refresh_from_db()
    assert document.etat == "processing"
    assert len(mailoutbox) == 0


def test_serveur_IA_envoie_l_etat_d_avancement_de_l_analyse_2(
    client, document, mailoutbox, alice
):
    rapport_csrd = document.rapports_csrd.select_related("entreprise").first()

    client.post(
        f"/ESRS-predict/{document.id}/{rapport_csrd.id}",
        {
            "status": "error",
            "msg": "MESSAGE",
        },
    )

    document.refresh_from_db()

    assert document.etat == "error"
    assert document.message == "MESSAGE"
    assert len(mailoutbox) == 1

    mail = mailoutbox[0]

    assert mail.from_email == settings.DEFAULT_FROM_EMAIL
    assert list(mail.to) == [alice.email]
    assert mail.template_id == settings.BREVO_RESULTAT_ANALYSE_IA_TEMPLATE


def test_serveur_IA_envoie_le_resultat_de_l_analyse(
    client, document, mailoutbox, alice
):
    rapport_csrd = document.rapports_csrd.select_related("entreprise").first()
    RESULTATS = """{
  "ESRS E1": [
    {
      "PAGES": 1,
      "TEXTS": "A"
    }
  ],
  "ESRS E2": [
    {
      "PAGES": 6,
      "TEXTS": "B"
    },
    {
      "PAGES": 7,
      "TEXTS": "C"
    }
  ]
  }"""

    response = client.post(
        f"/ESRS-predict/{document.id}/{rapport_csrd.id}",
        {
            "status": "success",
            "resultat_json": RESULTATS,
        },
    )

    document.refresh_from_db()
    assert document.etat == "success"
    assert document.resultat_json == RESULTATS

    assert len(mailoutbox) == 1
    mail = mailoutbox[0]
    assert mail.from_email == settings.DEFAULT_FROM_EMAIL
    assert list(mail.to) == [alice.email]
    assert mail.template_id == settings.BREVO_RESULTAT_ANALYSE_IA_TEMPLATE
    assert mail.merge_global_data == {
        "resultat_ia_url": response.wsgi_request.build_absolute_uri(
            reverse(
                "reglementations:gestion_csrd",
                kwargs={
                    "siren": rapport_csrd.entreprise.siren,
                    "id_etape": "analyse-ecart",
                },
            )
        )
        + "#onglets"
    }


def test_envoie_resultat_ia_email_non_bloquant(client, document, mocker):
    mocker.patch(
        "reglementations.views.csrd.analyse_ia.envoie_resultat_ia_email",
        side_effect=Exception,
    )
    capture_exception_mock = mocker.patch("sentry_sdk.capture_exception")

    rapport_csrd = document.rapports_csrd.select_related("entreprise").first()

    RESULTATS = """{
  "ESRS E1": [
    {
      "PAGES": 1,
      "TEXTS": "A"
    }
  ]
  }"""

    response = client.post(
        f"/ESRS-predict/{document.id}/{rapport_csrd.id}",
        {
            "status": "success",
            "resultat_json": RESULTATS,
        },
    )

    document.refresh_from_db()
    assert document.etat == "success"
    assert document.resultat_json == RESULTATS

    capture_exception_mock.assert_called_once()
    args, _ = capture_exception_mock.call_args
    assert isinstance(args[0], Exception)


def test_telechargement_des_resultats_IA_d_un_document_au_format_xlsx(
    client, csrd, alice
):
    document = AnalyseIA.objects.create(
        etat="success",
        resultat_json="""{
  "ESRS E1": [
    {
      "PAGES": 1,
      "TEXTS": "A"
    }
  ],
  "ESRS E2 : Pollution": [
    {
      "PAGES": 6,
      "TEXTS": "B"
    },
    {
      "PAGES": 7,
      "TEXTS": "C"
    }
  ]
  }""",
    )

    document.rapports_csrd.add(csrd)

    client.force_login(alice)

    response = client.get(
        f"/ESRS-predict/{document.id}/resultats.xlsx",
    )

    assert response["Content-Disposition"] == "filename=resultats.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook[">>>"]
    assert not onglet["C14"].value
    onglet = workbook["Phrases relatives aux ESRS"]
    assert onglet["A1"].value == "ESRS"
    assert onglet["B1"].value == "Fichier"
    assert onglet["C1"].value == "Page"
    assert onglet["D1"].value == "Phrase"
    assert onglet["A2"].value == "ESRS E1 - Changement climatique"
    assert onglet["C2"].value == 1
    assert onglet["D2"].value == "A"
    assert onglet["A3"].value == "ESRS E2 - Pollution"
    assert onglet["C3"].value == 6
    assert onglet["D3"].value == "B"
    assert onglet["A4"].value == "ESRS E2 - Pollution"
    assert onglet["C4"].value == 7
    assert onglet["D4"].value == "C"


def test_telechargement_des_resultats_IA_d_un_document_inexistant(client, alice):
    client.force_login(alice)

    response = client.get(
        "/ESRS-predict/42/resultats.xlsx",
    )

    assert response.status_code == 404


def test_telechargement_des_resultats_IA_d_un_document_redirige_vers_la_connexion_si_non_connecté(
    client, document
):
    response = client.get(
        f"/ESRS-predict/{document.id}/resultats.xlsx",
    )

    assert response.status_code == 302


def test_telechargement_des_resultats_ia_de_l_ensemble_des_documents_au_format_xlsx(
    client, csrd, alice
):
    document = AnalyseIA.objects.create(
        etat="success",
        resultat_json="""{
  "ESRS E1": [
    {
      "PAGES": 1,
      "TEXTS": "A"
    }
   ],
    "Non ESRS": [
    {
      "PAGES": 22,
      "TEXTS": "X"
    }
  ]
  }""",
    )
    document.rapports_csrd.add(csrd)

    document = AnalyseIA.objects.create(
        etat="success",
        resultat_json="""{
  "ESRS E2": [
    {
      "PAGES": 6,
      "TEXTS": "B"
    }
  ]
  }""",
    )
    document.rapports_csrd.add(csrd)

    client.force_login(alice)

    response = client.get(
        f"/ESRS-predict/{csrd.id}/synthese_resultats.xlsx",
    )

    assert response["Content-Disposition"] == "filename=synthese_resultats.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook[">>>"]
    assert onglet["C14"].value == "Synthèse générique"
    onglet = workbook["Phrases relatives aux ESRS"]
    assert onglet["A1"].value == "ESRS"
    assert onglet["B1"].value == "Fichier"
    assert onglet["C1"].value == "Page"
    assert onglet["D1"].value == "Phrase"
    assert onglet["A2"].value == "ESRS E1 - Changement climatique"
    assert onglet["C2"].value == 1
    assert onglet["D2"].value == "A"
    assert onglet["A3"].value == "ESRS E2 - Pollution"
    assert onglet["C3"].value == 6
    assert onglet["D3"].value == "B"


def test_telechargement_des_resultats_IA_de_l_ensemble_des_documents_d_un_rapport_csrd_inexistant(
    client, alice
):
    client.force_login(alice)

    response = client.get(
        "/ESRS-predict/42/synthese_resultats.xlsx",
    )

    assert response.status_code == 404


def test_telechargement_des_resultats_IA_de_l_ensemble_des_documents_redirige_vers_la_connexion_si_non_connecté(
    client, csrd
):
    response = client.get(
        f"/ESRS-predict/{csrd.id}/synthese_resultats.xlsx",
    )

    assert response.status_code == 302


def test_telechargement_des_resultats_par_ESRS_au_format_xlsx(client, csrd, alice):
    document = AnalyseIA.objects.create(
        etat="success",
        resultat_json="""{
  "ESRS E1": [
    {
      "PAGES": 1,
      "TEXTS": "A"
    }
  ],
  "ESRS E2": [
    {
      "PAGES": 4,
      "TEXTS": "B"
    }
  ],
  "Non ESRS": [
    {
      "PAGES": 22,
      "TEXTS": "X"
    }
  ]
  }""",
    )
    document.rapports_csrd.add(csrd)

    document = AnalyseIA.objects.create(
        etat="success",
        resultat_json="""{
  "ESRS E2": [
    {
      "PAGES": 6,
      "TEXTS": "C"
    }
  ]
  }""",
    )
    document.rapports_csrd.add(csrd)

    client.force_login(alice)

    response = client.get(
        f"/ESRS-predict/{csrd.id}/E2",
    )

    assert response["Content-Disposition"] == "filename=resultats_ESRS_E2.xlsx"
    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    onglet = workbook[">>>"]
    assert onglet["C14"].value == "ESRS E2 - Pollution"
    onglet = workbook["Phrases relatives aux ESRS"]
    assert onglet["A1"].value == "ESRS"
    assert onglet["B1"].value == "Fichier"
    assert onglet["C1"].value == "Page"
    assert onglet["D1"].value == "Phrase"
    assert onglet["A2"].value == "ESRS E2 - Pollution"
    assert onglet["C2"].value == 4
    assert onglet["D2"].value == "B"
    assert onglet["A3"].value == "ESRS E2 - Pollution"
    assert onglet["C3"].value == 6
    assert onglet["D3"].value == "C"


def test_telechargement_des_resultats_IA__par_ESRS_d_un_rapport_csrd_inexistant(
    client, alice
):
    client.force_login(alice)

    response = client.get(
        "/ESRS-predict/42/E2",
    )

    assert response.status_code == 404


def test_telechargement_des_resultats_IA_par_ESRS_redirige_vers_la_connexion_si_non_connecté(
    client, csrd
):
    response = client.get(
        f"/ESRS-predict/{csrd.id}/E2",
    )

    assert response.status_code == 302

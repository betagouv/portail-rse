from io import BytesIO

import pytest
from django.contrib.messages import WARNING
from django.urls import reverse
from openpyxl import load_workbook
from pytest_django.asserts import assertTemplateUsed

from habilitations.models import Habilitation
from reglementations.enums import EtapeCSRD
from reglementations.models.csrd import Enjeu
from reglementations.models.csrd import RapportCSRD


@pytest.mark.parametrize("etape", EtapeCSRD.ETAPES_VALIDABLES)
def test_gestion_de_la_csrd(etape, client, alice, entreprise_factory):
    entreprise = entreprise_factory()
    url = "/csrd/{siren}/etape-{etape}".format(siren=entreprise.siren, etape=etape)

    response = client.get(url)

    assert response.status_code == 302
    connexion_url = reverse("users:login")
    assert response.url == f"{connexion_url}?next={url}"

    Habilitation.ajouter(entreprise, alice, fonctions="Présidente")
    client.force_login(alice)
    response = client.get(url)

    assert response.status_code == 200
    context = response.context
    assert context["entreprise"] == entreprise
    if etape.endswith("introduction"):
        assertTemplateUsed(response, "reglementations/csrd/etape-introduction.html")
    elif etape.endswith("selection-enjeux"):
        assertTemplateUsed(response, "reglementations/csrd/etape-selection-enjeux.html")
    elif etape.endswith("analyse-materialite"):
        assertTemplateUsed(
            response, "reglementations/csrd/etape-analyse-materialite.html"
        )
    elif etape.endswith("selection-informations"):
        assertTemplateUsed(
            response, "reglementations/csrd/etape-selection-informations.html"
        )
    elif etape.endswith("analyse-ecart"):
        assertTemplateUsed(response, "reglementations/csrd/etape-analyse-ecart.html")
        assert context["synthese"] == {
            "phrases_environnement": [],
            "phrases_social": [],
            "phrases_gouvernance": [],
            "nb_phrases_pertinentes_detectees": 0,
            "nb_documents_analyses": 0,
            "nb_esrs_thematiques_detectees": 0,
        }
    elif etape.endswith("redaction-rapport-durabilite"):
        assertTemplateUsed(
            response, "reglementations/csrd/etape-redaction-rapport-durabilite.html"
        )

    rapport_csrd = RapportCSRD.objects.get(entreprise=entreprise)
    NOMBRE_ENJEUX = 103
    assert len(rapport_csrd.enjeux.all()) == NOMBRE_ENJEUX


def test_étape_inexistante_de_la_csrd(client, alice, entreprise_factory):
    entreprise = entreprise_factory()
    Habilitation.ajouter(entreprise, alice, fonctions="Présidente")
    client.force_login(alice)
    etape_inexistante = f"/csrd/{entreprise.siren}/etape-4"

    response = client.get(etape_inexistante)

    assert response.status_code == 404


ETAPES_ENREGISTRABLES = EtapeCSRD.ETAPES_VALIDABLES[:-1]


@pytest.mark.parametrize("etape", ETAPES_ENREGISTRABLES)
def test_enregistrement_de_l_étape_de_la_csrd(etape, client, csrd, alice):
    client.force_login(alice)
    url = "/csrd/{siren}/etape-{etape}".format(siren=csrd.entreprise.siren, etape=etape)

    client.post(url, follow=True)

    csrd.refresh_from_db()
    assert csrd.etape_validee == etape


@pytest.mark.parametrize("etape", ETAPES_ENREGISTRABLES)
def test_enregistrement_de_l_étape_de_la_csrd_retourne_une_404_si_aucune_CSRD(
    etape, client, alice, entreprise_factory
):
    entreprise = entreprise_factory()
    Habilitation.ajouter(entreprise, alice, fonctions="Présidente")
    client.force_login(alice)
    url = "/csrd/{siren}/etape-{etape}".format(siren=entreprise.siren, etape=etape)

    response = client.post(url, follow=True)

    assert response.status_code == 404
    assert RapportCSRD.objects.filter(entreprise=entreprise).count() == 0


def test_visualisation_des_enjeux(client, csrd, alice):
    client.force_login(alice)

    response = client.get(f"/csrd/fragments/selection_enjeux/{csrd.id}/ESRS_E1", {})

    assert response.status_code == 200
    assert "<!-- fragment enjeux -->" in response.content.decode("utf-8")


def test_selection_et_deselection_d_enjeux(client, csrd, alice):
    # update : les enjeux sont désormais sélectionnés par défaut
    enjeux = csrd.enjeux.all()
    enjeu_adaptation = enjeux[0]
    enjeu_attenuation = enjeux[1]
    enjeu_energie = enjeux[2]
    client.force_login(alice)

    response = client.post(
        f"/csrd/fragments/selection_enjeux/{csrd.id}/ESRS_E1",
        {"enjeux": [enjeu_adaptation.id, enjeu_energie.id]},
    )

    enjeu_adaptation.refresh_from_db()
    enjeu_attenuation.refresh_from_db()
    enjeu_energie.refresh_from_db()

    assert [enjeu_attenuation] == list(Enjeu.objects.filter(selection=False))
    assert response.status_code == 200

    context = response.context

    assert context["csrd"] == csrd
    assert "<!-- fragment esrs -->" in response.content.decode("utf-8")


def test_deselection_d_un_enjeu(client, csrd, alice):
    enjeux = csrd.enjeux.all()
    enjeu = enjeux[0]
    enjeu.selection = True
    enjeu.save()
    client.force_login(alice)

    response = client.post(
        f"/csrd/fragments/deselection_enjeu/{enjeu.id}",
    )

    enjeu.refresh_from_db()
    assert not enjeu.selection
    assert response.status_code == 200


def test_liste_des_enjeux_csrd_au_format_xlsx(client, csrd, alice):
    enjeux = csrd.enjeux.all()
    enjeu_adaptation = enjeux[0]
    enjeu_attenuation = enjeux[1]
    enjeu_attenuation.selection = True
    enjeu_attenuation.save()
    enjeu_energie = enjeux[2]
    client.force_login(alice)

    response = client.get(
        f"/csrd/{csrd.entreprise.siren}/enjeux.xlsx",
    )

    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )


def test_liste_des_enjeux_csrd__au_format_xlsx_retourne_une_404_si_entreprise_inexistante(
    client, alice
):
    client.force_login(alice)

    response = client.get(
        "/csrd/000000001/enjeux.xlsx",
    )

    assert response.status_code == 404


def test_liste_des_enjeux_csrd_au_format_xlsx_retourne_une_404_si_habilitation_inexistante(
    client, alice, entreprise_non_qualifiee
):
    client.force_login(alice)

    response = client.get(
        f"/csrd/{entreprise_non_qualifiee.siren}/enjeux.xlsx",
    )

    assert response.status_code == 404


def test_liste_des_enjeux_csrd_au_format_xlsx_retourne_une_404_si_csrd_inexistante(
    client, alice, entreprise_non_qualifiee
):
    Habilitation.ajouter(entreprise_non_qualifiee, alice, fonctions="Présidente")
    client.force_login(alice)

    response = client.get(
        f"/csrd/{entreprise_non_qualifiee.siren}/enjeux.xlsx",
    )

    assert response.status_code == 404


def test_liste_des_enjeux_csrd(client, csrd, alice):
    enjeux = csrd.enjeux.all()

    enjeu_attenuation = enjeux[1]
    enjeu_attenuation.selection = False
    enjeu_attenuation.save()

    assert (
        enjeux.filter(selection=False).count() == 1
    ), "Un des enjeux doit être désélectionné"

    client.force_login(alice)
    response = client.get(f"/csrd/fragments/liste_enjeux_selectionnes/{csrd.id}/1")

    assert response.status_code == 200

    context = response.context

    assert (
        len(
            context["enjeux_par_esg"]["environnement"][
                "ESRS E1 - Changement climatique"
            ]
        )
        == 2
    ), "Un des enjeux doit être désélectionné"
    assert "<!-- fragment liste des enjeux sélectionnés -->" in response.content.decode(
        "utf-8"
    )


def test_datapoints_pour_enjeux_materiels_au_format_xlsx(client, csrd, alice):
    enjeux = csrd.enjeux.all()
    enjeu_attenuation = enjeux[1]
    enjeu_attenuation.selection = True
    enjeu_attenuation.materiel = True
    enjeu_attenuation.save()
    esrs_materielle = enjeu_attenuation.esrs
    client.force_login(alice)

    response = client.get(
        f"/csrd/{csrd.entreprise.siren}/datapoints.xlsx",
    )

    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    noms_onglet = workbook.sheetnames
    assert esrs_materielle.replace("_", " ") in noms_onglet
    assert "Index" in noms_onglet
    assert "ESRS 2" in noms_onglet
    assert "ESRS2 MDR" in noms_onglet
    assert "ESRS G1" not in noms_onglet


def test_datapoints_pour_enjeux_non_materiels_au_format_xlsx(client, csrd, alice):
    enjeux = csrd.enjeux.all()

    # enjeu de l'ESRS_E1:
    enjeu_attenuation = enjeux[1]
    enjeu_attenuation.selection = True
    enjeu_attenuation.materiel = True
    enjeu_attenuation.save()
    esrs_materielle = enjeu_attenuation.esrs

    # note : les enjeux affichés dans le fichier "non-matériels"
    # doivent au préalable avoir été sélectionnés

    # enjeu de l'ESRS_G1:
    enjeux_G1 = enjeux.filter(esrs="ESRS_G1").first()
    enjeux_G1.selection = True
    enjeux_G1.materiel = False  # et pas None
    enjeux_G1.save()

    # les enjeux étant sélectionnés par défaut, on décoche certains ceux de l'ESRS E3
    # pour s'assurer que l'onglet de cet ESRS apparait bien dans le fichier des enjeux non-matériels.
    for enjeu_esrs_e3 in enjeux.filter(esrs="ESRS_E3"):
        enjeu_esrs_e3.selection = False
        enjeu_esrs_e3.save()

    client.force_login(alice)

    response = client.get(
        f"/csrd/{csrd.entreprise.siren}/datapoints.xlsx?materiel=false",
    )

    assert (
        response["content-type"]
        == "application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.content))
    noms_onglet = workbook.sheetnames

    assert esrs_materielle.replace("_", " ") not in noms_onglet
    assert "Index" in noms_onglet
    assert "ESRS G1" in noms_onglet
    assert "ESRS 2" not in noms_onglet
    assert "ESRS2 MDR" not in noms_onglet

    # Vérification de la présence des ESRS non-selectionnés
    assert (
        "ESRS E3" in noms_onglet
    ), "les enjeux non-sélectionnés doivent apparaitre dans le fichier"


def test_datapoints_csrd__au_format_xlsx_retourne_une_404_si_entreprise_inexistante(
    client, alice
):
    client.force_login(alice)

    response = client.get(
        "/csrd/000000001/datapoints.xlsx",
    )

    assert response.status_code == 404


def test_datapoints_csrd_au_format_xlsx_retourne_une_404_si_habilitation_inexistante(
    client, alice, entreprise_non_qualifiee
):
    client.force_login(alice)

    response = client.get(
        f"/csrd/{entreprise_non_qualifiee.siren}/datapoints.xlsx",
    )

    assert response.status_code == 404


def test_datapoints_csrd_au_format_xlsx_retourne_une_404_si_csrd_inexistante(
    client, alice, entreprise_non_qualifiee
):
    Habilitation.ajouter(entreprise_non_qualifiee, alice, fonctions="Présidente")
    client.force_login(alice)

    response = client.get(
        f"/csrd/{entreprise_non_qualifiee.siren}/datapoints.xlsx",
    )

    assert response.status_code == 404


def test_le_lien_analyse_d_écart_redirige_vers_l_étape_analyse_d_ecart_de_la_csrd(
    client, alice, entreprise_factory
):
    entreprise = entreprise_factory()
    Habilitation.ajouter(entreprise, alice, fonctions="Présidente")
    client.force_login(alice)
    url = "/csrd/etape-analyse-ecart"

    response = client.get(url, follow=True)

    assert response.status_code == 200
    redirect_url = reverse(
        "reglementations:gestion_csrd", args=[entreprise.siren, "analyse-ecart"]
    )
    assert response.redirect_chain == [(redirect_url, 302)]


def test_lien_analyse_d_écart_redirige_vers_la_page_d_ajout_d_entreprise_si_l_utilisateur_n_a_pas_d_entreprise(
    client, alice
):
    client.force_login(alice)
    url = "/csrd/etape-analyse-ecart"

    response = client.get(url, follow=True)

    assert response.status_code == 200
    assert response.redirect_chain == [(reverse("entreprises:entreprises"), 302)]
    messages = list(response.context["messages"])
    assert messages[0].level == WARNING
    assert (
        messages[0].message
        == "Commencez par ajouter une entreprise à votre compte utilisateur avant d'accéder à l'espace Rapport de Durabilité"
    )


def test_lien_analyse_d_écart_redirige_vers_la_page_de_connexion_si_l_utilisateur_n_est_pas_connecté(
    client,
):
    url = "/csrd/etape-analyse-ecart"

    response = client.get(url, follow=True)

    assert response.status_code == 200
    connexion_url = reverse("users:login")
    assert response.redirect_chain == [(f"{connexion_url}?next={url}", 302)]

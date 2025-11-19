from datetime import datetime
from datetime import timedelta
from functools import wraps
from pathlib import Path

from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist
from django.core.exceptions import PermissionDenied
from django.db.models import Q
from django.http import Http404
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.shortcuts import redirect
from django.template.loader import get_template
from django.template.loader import TemplateDoesNotExist
from django.urls import reverse_lazy
from openpyxl import load_workbook
from openpyxl import Workbook

from analyseia.forms import AnalyseIAForm
from analyseia.helpers import synthese_analyse
from entreprises.models import CaracteristiquesAnnuelles
from entreprises.models import Entreprise
from entreprises.views import get_current_entreprise
from habilitations.enums import UserRole
from habilitations.models import Habilitation
from reglementations.enums import ESRS
from reglementations.enums import EtapeCSRD
from reglementations.enums import ETAPES_CSRD
from reglementations.forms.csrd import LienRapportCSRDForm
from reglementations.models import RapportCSRD
from reglementations.views.base import Reglementation
from reglementations.views.base import ReglementationAction
from reglementations.views.base import ReglementationStatus
from utils.xlsx import xlsx_response


class CSRDReglementation(Reglementation):
    id = "csrd"
    title = "Rapport de Durabilité - CSRD"
    more_info_url = "https://portail-rse.beta.gouv.fr/fiches-reglementaires/rapport-de-durabilite-csrd/"
    tag = "tag-durabilite"
    summary = "Publier un rapport de durabilité."
    zone = "europe"

    @classmethod
    def est_suffisamment_qualifiee(cls, caracteristiques):
        return (
            caracteristiques.date_cloture_exercice is not None
            and caracteristiques.entreprise.est_cotee is not None
            and caracteristiques.entreprise.est_interet_public is not None
            and caracteristiques.effectif_securite_sociale is not None
            and caracteristiques.tranche_bilan is not None
            and caracteristiques.tranche_chiffre_affaires is not None
            and caracteristiques.entreprise.appartient_groupe is not None
            and (
                not caracteristiques.entreprise.appartient_groupe
                or (
                    caracteristiques.entreprise.comptes_consolides is not None
                    and (
                        not caracteristiques.entreprise.comptes_consolides
                        or (
                            caracteristiques.effectif_groupe is not None
                            and caracteristiques.tranche_bilan_consolide is not None
                            and caracteristiques.tranche_chiffre_affaires_consolide
                            is not None
                        )
                    )
                )
            )
        )

    @classmethod
    def est_soumis(cls, caracteristiques):
        super().est_soumis(caracteristiques)
        return bool(cls.est_soumis_a_partir_de_l_exercice(caracteristiques))

    @classmethod
    def est_soumis_a_partir_de_l_exercice(
        cls, caracteristiques: CaracteristiquesAnnuelles
    ) -> int | None:
        """Le premier exercice pour lequel l'entreprise est soumise à cette réglementation est l'exercice ouvert à compter du 1er janvier de l'année renvoyée"""
        if not (
            cls.critere_categorie_juridique_sirene(caracteristiques)
            or caracteristiques.entreprise.est_interet_public
        ):
            return
        if cls.est_grand_groupe(caracteristiques):
            if caracteristiques.entreprise.est_societe_mere:
                if (
                    caracteristiques.entreprise.est_cotee
                    or caracteristiques.entreprise.est_interet_public
                ) and caracteristiques.effectif_groupe in (
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_500_ET_4999,
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_5000_ET_9999,
                    CaracteristiquesAnnuelles.EFFECTIF_10000_ET_PLUS,
                ):
                    return 2024
                else:
                    return 2027
            else:
                if cls.est_grande_entreprise(caracteristiques):
                    if (
                        caracteristiques.entreprise.est_cotee
                        and caracteristiques.effectif_securite_sociale
                        == CaracteristiquesAnnuelles.EFFECTIF_SECURITE_SOCIALE_500_ET_PLUS
                    ):
                        return 2024
                    else:
                        return 2027
                elif (
                    caracteristiques.entreprise.est_cotee
                    and cls.est_petite_ou_moyenne_entreprise(caracteristiques)
                ):
                    if caracteristiques.effectif_groupe in (
                        CaracteristiquesAnnuelles.EFFECTIF_ENTRE_500_ET_4999,
                        CaracteristiquesAnnuelles.EFFECTIF_ENTRE_5000_ET_9999,
                        CaracteristiquesAnnuelles.EFFECTIF_10000_ET_PLUS,
                    ):
                        return 2024
                    else:
                        return 2027
        else:
            if caracteristiques.entreprise.est_dans_EEE:
                if caracteristiques.entreprise.est_cotee:
                    if cls.est_grande_entreprise(caracteristiques):
                        if (
                            caracteristiques.effectif_securite_sociale
                            == CaracteristiquesAnnuelles.EFFECTIF_SECURITE_SOCIALE_500_ET_PLUS
                        ):
                            return 2024
                        else:
                            return 2027
                    elif cls.est_petite_ou_moyenne_entreprise(caracteristiques):
                        return 2028
                elif caracteristiques.entreprise.est_interet_public:
                    if cls.est_grande_entreprise(caracteristiques):
                        if (
                            caracteristiques.effectif_securite_sociale
                            == CaracteristiquesAnnuelles.EFFECTIF_SECURITE_SOCIALE_500_ET_PLUS
                        ):
                            return 2024
                        else:
                            return 2027
                elif cls.est_grande_entreprise(caracteristiques):
                    return 2027
            else:
                if cls.est_grande_entreprise(caracteristiques):
                    return 2027
                elif cls.est_micro_ou_petite_entreprise_hors_EEE_consideree_comme_soumise(
                    caracteristiques
                ):
                    return 2028

    @classmethod
    def criteres_remplis(cls, caracteristiques):
        criteres = []
        if (
            caracteristiques.entreprise.est_hors_EEE
            and not caracteristiques.entreprise.appartient_groupe
        ):
            criteres.append("votre siège social est hors EEE")
        else:
            if caracteristiques.entreprise.est_cotee:
                criteres.append("votre société est cotée sur un marché réglementé")
            if caracteristiques.entreprise.est_interet_public:
                criteres.append("votre société est d'intérêt public")
            if caracteristiques.entreprise.est_societe_mere and cls.est_grand_groupe(
                caracteristiques
            ):
                criteres.append("votre société est la société mère d'un groupe")
        if critere := cls.critere_effectif(caracteristiques):
            criteres.append(critere)
        if critere := cls.critere_bilan(caracteristiques):
            criteres.append(critere)
        if critere := cls.critere_CA(caracteristiques):
            criteres.append(critere)
        return criteres

    @classmethod
    def critere_effectif(cls, caracteristiques):
        if (
            caracteristiques.entreprise.est_cotee
            or caracteristiques.entreprise.est_interet_public
        ):
            if cls.est_grande_entreprise(caracteristiques):
                if (
                    caracteristiques.effectif_securite_sociale
                    == CaracteristiquesAnnuelles.EFFECTIF_SECURITE_SOCIALE_500_ET_PLUS
                ):
                    return "votre effectif est supérieur à 500 salariés"
                elif (
                    caracteristiques.effectif_securite_sociale
                    == CaracteristiquesAnnuelles.EFFECTIF_SECURITE_SOCIALE_ENTRE_250_ET_499
                ):
                    return "votre effectif est supérieur à 250 salariés"
            if cls.est_grand_groupe(caracteristiques):
                if caracteristiques.effectif_groupe in (
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_500_ET_4999,
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_5000_ET_9999,
                    CaracteristiquesAnnuelles.EFFECTIF_10000_ET_PLUS,
                ):
                    return "l'effectif du groupe est supérieur à 500 salariés"
                elif (
                    caracteristiques.effectif_groupe
                    == CaracteristiquesAnnuelles.EFFECTIF_ENTRE_250_ET_499
                ):
                    return "l'effectif du groupe est supérieur à 250 salariés"
            if cls.est_petite_ou_moyenne_entreprise(caracteristiques):
                if (
                    caracteristiques.effectif_securite_sociale
                    != CaracteristiquesAnnuelles.EFFECTIF_SECURITE_SOCIALE_MOINS_DE_10
                ):
                    return "votre effectif est supérieur à 10 salariés"
        else:
            if cls.est_grande_entreprise(caracteristiques):
                if caracteristiques.effectif_securite_sociale in (
                    CaracteristiquesAnnuelles.EFFECTIF_SECURITE_SOCIALE_ENTRE_250_ET_499,
                    CaracteristiquesAnnuelles.EFFECTIF_SECURITE_SOCIALE_500_ET_PLUS,
                ):
                    return "votre effectif est supérieur à 250 salariés"
            if caracteristiques.entreprise.est_societe_mere and cls.est_grand_groupe(
                caracteristiques
            ):
                if caracteristiques.effectif_groupe in (
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_250_ET_499,
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_500_ET_4999,
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_5000_ET_9999,
                    CaracteristiquesAnnuelles.EFFECTIF_10000_ET_PLUS,
                ):
                    return "l'effectif du groupe est supérieur à 250 salariés"

    @classmethod
    def critere_bilan(cls, caracteristiques):
        if cls.est_grand_groupe(caracteristiques):
            if caracteristiques.tranche_bilan_consolide in (
                CaracteristiquesAnnuelles.BILAN_ENTRE_30M_ET_43M,
                CaracteristiquesAnnuelles.BILAN_ENTRE_43M_ET_100M,
                CaracteristiquesAnnuelles.BILAN_100M_ET_PLUS,
            ):
                return "le bilan du groupe est supérieur à 30M€"
        if caracteristiques.tranche_bilan in (
            CaracteristiquesAnnuelles.BILAN_ENTRE_25M_ET_43M,
            CaracteristiquesAnnuelles.BILAN_ENTRE_43M_ET_100M,
            CaracteristiquesAnnuelles.BILAN_100M_ET_PLUS,
        ):
            if cls.est_grande_entreprise(caracteristiques):
                return "votre bilan est supérieur à 25M€"
            elif cls.est_petite_ou_moyenne_entreprise(caracteristiques):
                return "votre bilan est supérieur à 450k€"
        elif (
            caracteristiques.tranche_bilan
            == CaracteristiquesAnnuelles.BILAN_ENTRE_450K_ET_25M
        ):
            if cls.est_petite_ou_moyenne_entreprise(caracteristiques):
                return "votre bilan est supérieur à 450k€"

    @classmethod
    def critere_CA(cls, caracteristiques):
        if cls.est_grand_groupe(caracteristiques):
            if caracteristiques.tranche_chiffre_affaires_consolide in (
                CaracteristiquesAnnuelles.CA_ENTRE_60M_ET_100M,
                CaracteristiquesAnnuelles.CA_100M_ET_PLUS,
            ):
                return "le chiffre d'affaires du groupe est supérieur à 60M€"
        if caracteristiques.tranche_chiffre_affaires in (
            CaracteristiquesAnnuelles.CA_ENTRE_50M_ET_100M,
            CaracteristiquesAnnuelles.CA_100M_ET_PLUS,
        ):
            if cls.est_grande_entreprise(caracteristiques):
                return "votre chiffre d'affaires est supérieur à 50M€"
            elif cls.est_petite_ou_moyenne_entreprise(caracteristiques):
                return "votre chiffre d'affaires est supérieur à 900k€"
        elif (
            caracteristiques.tranche_chiffre_affaires
            == CaracteristiquesAnnuelles.CA_ENTRE_900K_ET_50M
        ):
            if cls.est_petite_ou_moyenne_entreprise(caracteristiques):
                return "votre chiffre d'affaires est supérieur à 900k€"

    @classmethod
    def critere_categorie_juridique_sirene(cls, caracteristiques):
        return (
            caracteristiques.entreprise.categorie_juridique_sirene == 3120
            or 5100 <= caracteristiques.entreprise.categorie_juridique_sirene <= 6199
            or 6300 <= caracteristiques.entreprise.categorie_juridique_sirene <= 6499
            or 8100 <= caracteristiques.entreprise.categorie_juridique_sirene <= 8299
        )

    @classmethod
    def est_delegable(cls, caracteristiques):
        if caracteristiques.entreprise.est_societe_mere:
            return False
        elif (
            cls.est_grande_entreprise(caracteristiques)
            and caracteristiques.entreprise.est_cotee
        ):
            return False
        return cls.est_grand_groupe(caracteristiques)

    @classmethod
    def calculate_status(
        cls,
        caracteristiques: CaracteristiquesAnnuelles,
    ) -> ReglementationStatus:
        if reglementation_status := super().calculate_status(caracteristiques):
            return reglementation_status

        try:
            rapport = rapport_csrd(
                entreprise=caracteristiques.entreprise,
                annee=datetime.today().year,
            )
        except ObjectDoesNotExist:
            rapport = None

        if rapport and rapport.etape_validee:
            etape_suivante = EtapeCSRD.id_suivant(rapport.etape_validee)
            label_gestion_csrd = "Reprendre l’actualisation de mon rapport"
        else:
            etape_suivante = "introduction"
            label_gestion_csrd = "Actualiser mon Rapport de Durabilité"

        # on considère que la publication du rapport est la dernière étape
        rapport_termine = rapport and rapport.est_termine

        if rapport_termine:
            statut_courant = ReglementationStatus.STATUS_A_JOUR
        elif not rapport:
            statut_courant = ReglementationStatus.STATUS_A_ACTUALISER
        else:
            statut_courant = ReglementationStatus.STATUS_EN_COURS

        primary_action = ReglementationAction(
            reverse_lazy(
                "reglementations:gestion_csrd",
                kwargs={
                    "siren": caracteristiques.entreprise.siren,
                    "id_etape": etape_suivante,
                },
            ),
            label_gestion_csrd,
        )
        if annee := cls.est_soumis_a_partir_de_l_exercice(caracteristiques):
            premiere_annee_publication = (
                cls.decale_annee_publication_selon_cloture_exercice_comptable(
                    annee, caracteristiques
                )
            )
            exercice_comptable = (
                f"{annee}"
                if caracteristiques.exercice_comptable_est_annee_civile
                else f"{annee}-{annee + 1}"
            )
            status_detail = f"Vous êtes soumis à cette réglementation à partir de {premiere_annee_publication} sur les données de l'exercice comptable {exercice_comptable}"
            if cls.est_micro_ou_petite_entreprise_hors_EEE_consideree_comme_soumise(
                caracteristiques
            ):
                conditions = "votre société dont le siège social est hors EEE revêt une forme juridique comparable aux sociétés par actions ou aux sociétés à responsabilité limitée, comptabilise un chiffre d'affaires net dans l'Espace économique européen qui excède 150 millions d'euros à la date de clôture des deux derniers exercices consécutifs, ne contrôle ni n'est contrôlée par une autre société et dispose d'une succursale en France dont le chiffre d'affaires net excède 40 millions d'euros"
                status_detail += f" si {conditions}."
            else:
                criteres = cls.criteres_remplis(caracteristiques)
                justification = ", ".join(criteres[:-1]) + " et " + criteres[-1]
                status_detail += f" car {justification}."
                if cls.est_delegable(caracteristiques):
                    status_detail += (
                        " Vous pouvez déléguer cette obligation à votre société-mère."
                    )
            status_detail += " Vous devez publier le Rapport de Durabilité en même temps que le rapport de gestion."
            # notes (FV):
            # AMA, ces aspects sont métiers et devraient êtres rattaché à l'objet métier correspondant
            # c.a.d `RapportCSRD`
            return ReglementationStatus(
                status=statut_courant,
                status_detail=status_detail,
                primary_action=primary_action,
                prochaine_echeance=premiere_annee_publication,
            )
        else:
            primary_action.title = "Tester un Rapport de Durabilité"
            return ReglementationStatus(
                status=ReglementationStatus.STATUS_NON_SOUMIS,
                status_detail="Vous n'êtes pas soumis à cette réglementation.",
                primary_action=primary_action,
            )

    @staticmethod
    def decale_annee_publication_selon_cloture_exercice_comptable(
        annee, caracteristiques
    ):
        ouverture_exercice_comptable = (
            caracteristiques.date_cloture_exercice + timedelta(days=1)
        ) + relativedelta(
            year=annee
        )  # remplace l'année de la date d'ouverture de l'exercice comptable par la valeur donnée en paramètre (de manière intelligente pour les années bissextiles)
        cloture_exercice_comptable = (
            ouverture_exercice_comptable + relativedelta(months=+12) - timedelta(days=1)
        )
        return (cloture_exercice_comptable + relativedelta(months=+6)).year

    @classmethod
    def est_microentreprise(cls, caracteristiques: CaracteristiquesAnnuelles):
        nombre_seuils_non_depasses = 0
        if (
            caracteristiques.tranche_bilan
            == CaracteristiquesAnnuelles.BILAN_MOINS_DE_450K
        ):
            nombre_seuils_non_depasses += 1
        if (
            caracteristiques.tranche_chiffre_affaires
            == CaracteristiquesAnnuelles.CA_MOINS_DE_900K
        ):
            nombre_seuils_non_depasses += 1
        if (
            caracteristiques.effectif_securite_sociale
            == CaracteristiquesAnnuelles.EFFECTIF_SECURITE_SOCIALE_MOINS_DE_10
        ):
            nombre_seuils_non_depasses += 1
        return nombre_seuils_non_depasses >= 2

    @classmethod
    def est_grande_entreprise(cls, caracteristiques: CaracteristiquesAnnuelles) -> bool:
        nombre_seuils_depasses = 0
        if caracteristiques.tranche_bilan in (
            CaracteristiquesAnnuelles.BILAN_ENTRE_25M_ET_43M,
            CaracteristiquesAnnuelles.BILAN_ENTRE_43M_ET_100M,
            CaracteristiquesAnnuelles.BILAN_100M_ET_PLUS,
        ):
            nombre_seuils_depasses += 1
        if caracteristiques.tranche_chiffre_affaires in (
            CaracteristiquesAnnuelles.CA_ENTRE_50M_ET_100M,
            CaracteristiquesAnnuelles.CA_100M_ET_PLUS,
        ):
            nombre_seuils_depasses += 1
        if caracteristiques.effectif_securite_sociale in (
            CaracteristiquesAnnuelles.EFFECTIF_SECURITE_SOCIALE_ENTRE_250_ET_499,
            CaracteristiquesAnnuelles.EFFECTIF_SECURITE_SOCIALE_500_ET_PLUS,
        ):
            nombre_seuils_depasses += 1
        return nombre_seuils_depasses >= 2

    @classmethod
    def est_petite_ou_moyenne_entreprise(
        cls, caracteristiques: CaracteristiquesAnnuelles
    ) -> bool:
        return not cls.est_microentreprise(
            caracteristiques
        ) and not cls.est_grande_entreprise(caracteristiques)

    @classmethod
    def est_grand_groupe(cls, caracteristiques: CaracteristiquesAnnuelles) -> bool:
        nombre_seuils_depasses = 0
        if caracteristiques.tranche_bilan_consolide in (
            CaracteristiquesAnnuelles.BILAN_ENTRE_30M_ET_43M,
            CaracteristiquesAnnuelles.BILAN_ENTRE_43M_ET_100M,
            CaracteristiquesAnnuelles.BILAN_100M_ET_PLUS,
        ):
            nombre_seuils_depasses += 1
        if caracteristiques.tranche_chiffre_affaires_consolide in (
            CaracteristiquesAnnuelles.CA_ENTRE_60M_ET_100M,
            CaracteristiquesAnnuelles.CA_100M_ET_PLUS,
        ):
            nombre_seuils_depasses += 1
        if caracteristiques.effectif_groupe in (
            CaracteristiquesAnnuelles.EFFECTIF_ENTRE_250_ET_499,
            CaracteristiquesAnnuelles.EFFECTIF_ENTRE_500_ET_4999,
            CaracteristiquesAnnuelles.EFFECTIF_ENTRE_5000_ET_9999,
            CaracteristiquesAnnuelles.EFFECTIF_10000_ET_PLUS,
        ):
            nombre_seuils_depasses += 1
        return (
            caracteristiques.entreprise.appartient_groupe
            and nombre_seuils_depasses >= 2
        )

    @classmethod
    def est_micro_ou_petite_entreprise_hors_EEE_consideree_comme_soumise(
        cls, caracteristiques: CaracteristiquesAnnuelles
    ) -> bool:
        """
        Si une entreprise hors EEE est une grande entreprise ou fait partie d'un grand groupe elle est soumise à la CSRD comme les entreprise EEE.
        Sinon, elle n'est soumise que dans les conditions cumulatives suivantes :
        - Revêt une forme juridique comparable aux sociétés par actions et aux sociétés à responsabilité limitée ;
        - Comptabilise un chiffre d'affaires net dans l'Espace économique européen qui excède 150 millions d'euros, à la date de clôture des deux derniers exercices consécutifs ;
        - Ne contrôle ni n'est contrôlée par une autre société
        - Dispose d'une succursale en France dont le chiffre d'affaires net excède 40 millions d'euros
        On approxime ce cas très à la marge sur le portail en considérant qu'elle est soumise dès lors que son CA est supérieur à 100M€ (il n'y a pas de seuil à 150M€ dans les tranches actuelles)
        mais on précisera les conditions exactes dans le détail de l'explication donnée à l'utilisateur.
        """
        return (
            not caracteristiques.entreprise.est_dans_EEE
            and not cls.est_grand_groupe(caracteristiques)
            and not cls.est_grande_entreprise(caracteristiques)
            and caracteristiques.tranche_chiffre_affaires
            == CaracteristiquesAnnuelles.CA_100M_ET_PLUS
        )


def rapport_csrd(entreprise, annee):
    """Cherche un RapportCSRD

    Lève une exception si le RapportCSRD n'existe pas
    """
    return RapportCSRD.objects.get(
        entreprise=entreprise,
        proprietaire=None,  # à supprimer une fois les documents personnels supprimés de la bdd
        annee=annee,
    )


@login_required
def gestion_csrd(request, siren=None, id_etape="introduction"):
    if not siren:
        entreprise = get_current_entreprise(request)
        if not entreprise:
            messages.warning(
                request,
                "Commencez par ajouter une entreprise à votre compte utilisateur avant d'accéder à l'espace Rapport de Durabilité",
            )
            return redirect("entreprises:entreprises")
        return redirect(
            "reglementations:gestion_csrd", siren=entreprise.siren, id_etape=id_etape
        )

    entreprise = get_object_or_404(Entreprise, siren=siren)
    if not Habilitation.existe(entreprise, request.user):
        raise PermissionDenied

    template_name = f"reglementations/csrd/etape-{id_etape}.html"
    try:
        template = get_template(template_name)
    except TemplateDoesNotExist:
        raise Http404

    # En analysant un peu les requêtes exécutées,
    # les fetch sur les habilitations et l'entreprise se répètent (requêtes identiques)
    # un peu de centralisation à faire pour éviter les répétitions dans `utils.middlewares.ExtendUserMiddleware`

    role = Habilitation.role_pour(entreprise, request.user)
    annee = datetime.now().year

    if request.method == "POST":
        try:
            csrd = rapport_csrd(entreprise, annee)
        except ObjectDoesNotExist:
            raise Http404
        csrd.etape_validee = id_etape
        csrd.save()
        return redirect(
            "reglementations:gestion_csrd",
            siren=siren,
            id_etape=EtapeCSRD.id_suivant(id_etape),
        )

    # on vérifie si on a sélectionné une année manuellement
    if csrd_id := request.session.get("rapport_csrd_courant"):
        try:
            csrd = RapportCSRD.objects.prefetch_related("enjeux", "enjeux__parent").get(
                pk=csrd_id, entreprise=entreprise
            )
        except RapportCSRD.DoesNotExist:
            # par ex. : l'utilisateur a selectionné une autre entreprise
            request.session.pop("rapport_csrd_courant")

    # la suppression de la notion de propriétaire implique une gestion des droits
    # par ex. un lecteur ne pourra pas créer de rapport CSRD
    if not request.session.get("rapport_csrd_courant"):
        # un lecteur ou un éditeur ne peuvent pas créer de rapport CSRD
        if role == UserRole.PROPRIETAIRE:
            # les prefetch de l'enjeu parent évitent des N+1 au niveau du template
            csrd, _ = RapportCSRD.objects.prefetch_related(
                "enjeux", "enjeux__parent"
            ).get_or_create(
                entreprise=entreprise,
                proprietaire=None,  # à supprimer une fois les documents personnels supprimés de la bdd
                annee=annee,
            )
        else:
            # si l'utilisateur n'est pas propriétaire, il ne peut que consulter
            # ou modifier un rapport existant
            csrd = RapportCSRD.objects.filter(
                entreprise=entreprise,
                proprietaire=None,  # à supprimer une fois les documents personnels supprimés de la bdd
                annee=annee,
            )
            if not csrd.exists():
                # pas de rapport existant créé par un propriétaire :
                # on notifie et on bloque l'accès
                messages.warning(
                    request,
                    "Il n'y a pas encore de rapport CSRD créé pour cette année (possible uniquement en ayant le rôle propriétaire).",
                )
                return redirect("reglementations:tableau_de_bord", siren=siren)

            # on récupére le rapport officiel,
            # la gestion des droits de modification éventuels est au niveau du template
            csrd = csrd.first()

    context = contexte_d_etape(id_etape, csrd)
    return HttpResponse(template.render(context, request))


def contexte_d_etape(id_etape, csrd, form=None):
    # on récupère les rapports si plusieurs existants, pour permettre une sélection
    rapports_csrd = RapportCSRD.objects.filter(
        entreprise=csrd.entreprise,
        proprietaire=csrd.proprietaire,  # à supprimer une fois les documents personnels supprimés de la bdd
    ).order_by("annee")

    context = {
        "entreprise": csrd.entreprise,
        "reglementation": CSRDReglementation,
        "etape": EtapeCSRD.get(id_etape),
        "csrd": csrd,
        "annee": datetime.now().year,
        "steps": ETAPES_CSRD,
        "rapports_csrd": rapports_csrd,
    }

    match EtapeCSRD.get(id_etape).id:
        ## légèrement plus lisible qu'un `if`
        case "selection-informations":
            nb_enjeux_non_analyses = csrd.enjeux.non_analyses().count()
            context |= {
                "can_download": nb_enjeux_non_analyses
                != csrd.enjeux.selectionnes().count(),
                "nb_enjeux_non_analyses": nb_enjeux_non_analyses,
            }
        case "analyse-ecart":
            context |= {
                "form": form or AnalyseIAForm(),
                "analyses_ia": csrd.analyses_ia.all(),
                "synthese": synthese_analyse(
                    csrd.analyses_ia.reussies(), prefixe_ESRS=True
                ),
                "onglet_resultats_actif": csrd.analyses_ia.reussies().exists()
                and not csrd.analyses_ia.non_lancees(),
            }
        case "redaction-rapport-durabilite":
            context |= {"form": LienRapportCSRDForm(instance=csrd)}

    return context


def csrd_required_with_enjeux(function):
    @wraps(function)
    def wrap(request, siren):
        entreprise = get_object_or_404(Entreprise, siren=siren)
        try:
            habilitation = request.user.habilitation_set.get(entreprise=entreprise)
            csrd = RapportCSRD.objects.prefetch_related("enjeux", "enjeux__parent").get(
                entreprise=entreprise,
                proprietaire=None,  # à supprimer une fois les documents personnels supprimés de la bdd
                annee=datetime.now().year,
            )
            return function(request, siren, csrd=csrd)
        except ObjectDoesNotExist:
            raise Http404("Ce rapport de durabilité n'existe pas")

    return wrap


@login_required
@csrd_required_with_enjeux
def enjeux_xlsx(request, siren, csrd=None):
    return _build_xlsx(csrd.enjeux, csrd=csrd)


@login_required
@csrd_required_with_enjeux
def enjeux_materiels_xlsx(request, siren, csrd=None):
    return _build_xlsx(csrd.enjeux, csrd=csrd, materiels=True)


def _build_xlsx(enjeux, csrd=None, materiels=False):
    # construit le fichier XLS pour :
    # - les enjeux sélectionnés
    # - les enjeux analysés (matériels ou non)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "ESRS"
    worksheet["B1"] = "Enjeux de durabilité"
    worksheet["C1"] = "Descriptif"
    worksheet["D1"] = "Origine"

    if materiels:
        worksheet["E1"] = "Materialité"

    numero_ligne = 2
    for esrs in ESRS:
        enjeux = csrd.enjeux_par_esrs(esrs)

        if materiels:
            enjeux = enjeux.filter(materiel__isnull=False)

        for enjeu in enjeux:
            if enjeu.selection:
                worksheet[f"A{numero_ligne}"] = enjeu.esrs
                worksheet[f"B{numero_ligne}"] = enjeu.nom
                worksheet[f"C{numero_ligne}"] = enjeu.description
                worksheet[f"D{numero_ligne}"] = (
                    "personnel" if enjeu.modifiable else "AR"
                )

                if materiels:
                    worksheet[f"E{numero_ligne}"] = (
                        "Matériel" if enjeu.materiel else "Non-matériel"
                    )

                numero_ligne += 1

    filename = "enjeux_csrd.xlsx" if not materiels else "enjeux_csrd_materiels.xlsx"
    return xlsx_response(workbook, filename)


@login_required
@csrd_required_with_enjeux
def datapoints_xlsx(request, _, csrd=None):
    materiel = request.GET.get("materiel", True) != "false"
    esrs_a_supprimer = _esrs_materiel_a_supprimer(csrd, materiel)
    chemin_xlsx = Path(settings.BASE_DIR, "static/CSRD/ESRS_Data_Points_EFRAG.xlsx")
    workbook = load_workbook(chemin_xlsx)
    for esrs in esrs_a_supprimer:
        titre_onglet = esrs.replace("_", " ")
        workbook.remove(workbook[titre_onglet])

    filename = (
        "datapoints_csrd_materiels.xlsx"
        if materiel
        else "datapoints_csrd_non_materiels.xlsx"
    )
    return xlsx_response(workbook, filename)


def _esrs_materiel_a_supprimer(csrd: RapportCSRD, materiel: bool):
    tous_les_esrs = set(ESRS.values)

    if materiel:
        enjeux_materiels = csrd.enjeux.materiels()
        esrs_materiels = set((enjeu.esrs for enjeu in enjeux_materiels))
        esrs_a_supprimer = tous_les_esrs - esrs_materiels
        esrs_a_supprimer.discard("ESRS_2")
    else:
        # les enjeux non-analysés doivent aussi être dans ce document (comme non-matériels)
        enjeux_non_materiels = csrd.enjeux.filter(
            Q(materiel=False) | Q(selection=False)
        )
        esrs_non_materiels = set((enjeu.esrs for enjeu in enjeux_non_materiels))
        # pour les enjeux non-matériels, ESRS_2 et ESRS_2_MDR sont à supprimer
        esrs_a_supprimer = tous_les_esrs - esrs_non_materiels
        esrs_a_supprimer |= {"ESRS2_MDR", "ESRS_2"}

    # ne pas supprimer ESRS_1 car il n'existe pas dans le fichier .xlsx
    esrs_a_supprimer.discard("ESRS_1")

    return esrs_a_supprimer

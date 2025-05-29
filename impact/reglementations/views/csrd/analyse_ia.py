import json
from pathlib import Path

import sentry_sdk
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist
from django.core.mail import EmailMessage
from django.http import Http404
from django.http import HttpResponse
from django.shortcuts import redirect
from django.shortcuts import render
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from openpyxl import load_workbook

from api import analyse_ia
from api.exceptions import APIError
from reglementations.forms.csrd import DocumentAnalyseIAForm
from reglementations.models import DocumentAnalyseIA
from reglementations.models import RapportCSRD
from reglementations.views.csrd.csrd import contexte_d_etape
from reglementations.views.csrd.csrd import normalise_titre_esrs
from reglementations.views.csrd.csrd import xlsx_response
from reglementations.views.csrd.decorators import csrd_required
from reglementations.views.csrd.decorators import document_required


@login_required
@csrd_required
@require_http_methods(["POST"])
def ajout_document(request, csrd_id):
    id_etape = "analyse-ecart"
    csrd = RapportCSRD.objects.get(id=csrd_id)
    data = {**request.POST}
    data["rapport_csrd"] = csrd_id
    form = DocumentAnalyseIAForm(data=data, files=request.FILES)
    if form.is_valid():
        form.save()
        messages.success(request, "Document ajouté")
        return redirect(
            "reglementations:gestion_csrd",
            siren=csrd.entreprise.siren,
            id_etape=id_etape,
        )
    else:
        context = contexte_d_etape(id_etape, csrd, form)
        template_name = f"reglementations/csrd/etape-{id_etape}.html"
        return render(request, template_name, context, status=400)


@login_required
@document_required
@require_http_methods(["POST"])
def suppression_document(request, id_document):
    document = DocumentAnalyseIA.objects.get(id=id_document)
    document.delete()
    document.fichier.delete(save=False)
    messages.success(request, "Document supprimé")
    id_etape = "analyse-ecart"
    return redirect(
        "reglementations:gestion_csrd",
        siren=document.rapport_csrd.entreprise.siren,
        id_etape=id_etape,
    )


@login_required
@document_required
@require_http_methods(["POST"])
def lancement_analyse_IA(request, id_document):
    document = DocumentAnalyseIA.objects.get(id=id_document)

    if document.etat != "success":
        try:
            callback_url = request.build_absolute_uri(
                reverse(
                    "reglementations:etat_analyse_IA",
                    kwargs={
                        "id_document": document.id,
                    },
                )
            )
            etat = analyse_ia.lancement_analyse(
                document.id, document.fichier.url, callback_url
            )
            document.etat = etat
            document.save()
            messages.success(
                request,
                "L'analyse a bien été lancée. Celle-ci peut prendre entre quelques secondes et quelques heures. Vous recevrez un email lorsque les résultats seront disponibles.",
            )
        except APIError as exception:
            messages.error(request, exception)

    id_etape = "analyse-ecart"
    return redirect(
        "reglementations:gestion_csrd",
        siren=document.rapport_csrd.entreprise.siren,
        id_etape=id_etape,
    )


@csrf_exempt
def etat_analyse_IA(request, id_document):
    try:
        document = DocumentAnalyseIA.objects.get(id=id_document)
    except ObjectDoesNotExist:
        raise Http404("Ce document n'existe pas")

    status = request.POST.get("status")
    document.etat = status
    if message := request.POST.get("msg"):
        document.message = message
    if status == "success":
        document.resultat_json = request.POST["resultat_json"]
    document.save()
    if status in ("success", "error"):
        path = reverse(
            "reglementations:gestion_csrd",
            kwargs={
                "siren": document.rapport_csrd.entreprise.siren,
                "id_etape": "analyse-ecart",
            },
        )
        try:
            envoie_resultat_ia_email(
                document, f"{request.build_absolute_uri(path)}#onglets"
            )
        except Exception as e:
            with sentry_sdk.new_scope() as scope:
                scope.set_level("info")
                sentry_sdk.capture_exception(e)

    return HttpResponse("OK")


def envoie_resultat_ia_email(document, resultat_ia_url):
    destinataires = [
        utilisateur.email
        for utilisateur in document.rapport_csrd.entreprise.users.all()
    ]

    email = EmailMessage(
        to=destinataires,
        from_email=settings.DEFAULT_FROM_EMAIL,
    )
    email.template_id = settings.BREVO_RESULTAT_ANALYSE_IA_TEMPLATE

    email.merge_global_data = {
        "resultat_ia_url": resultat_ia_url,
    }
    email.send()


@login_required
@document_required
def resultat_IA_xlsx(request, id_document):
    document = DocumentAnalyseIA.objects.get(id=id_document)

    chemin_xlsx = Path(
        settings.BASE_DIR, "reglementations/views/csrd/xlsx/template_synthese_ESG.xlsx"
    )
    workbook = load_workbook(chemin_xlsx)
    worksheet = workbook[">>>"]
    worksheet["C14"] = ""
    worksheet = workbook["Phrases relatives aux ESRS"]
    _ajoute_ligne_resultat_ia(worksheet, document, True, None)
    return xlsx_response(workbook, "resultats.xlsx")


def _ajoute_ligne_resultat_ia(worksheet, document, avec_nom_fichier, contrainte_esrs):
    data = json.loads(document.resultat_json)
    for esrs, contenus in data.items():
        for contenu in contenus:
            if esrs != "Non ESRS" and (
                (not contrainte_esrs) or contrainte_esrs in esrs
            ):
                if avec_nom_fichier:
                    ligne = [
                        normalise_titre_esrs(esrs),
                        document.nom,
                        contenu["PAGES"],
                        contenu["TEXTS"],
                    ]
                else:
                    ligne = [
                        normalise_titre_esrs(esrs),
                        contenu["PAGES"],
                        contenu["TEXTS"],
                    ]
                worksheet.append(ligne)


@login_required
@csrd_required
def synthese_resultat_IA_xlsx(request, csrd_id):
    csrd = RapportCSRD.objects.get(id=csrd_id)
    chemin_xlsx = Path(
        settings.BASE_DIR, "reglementations/views/csrd/xlsx/template_synthese_ESG.xlsx"
    )
    workbook = load_workbook(chemin_xlsx)
    worksheet = workbook["Phrases relatives aux ESRS"]
    for document in csrd.documents_analyses:
        _ajoute_ligne_resultat_ia(worksheet, document, True, None)
    return xlsx_response(workbook, "synthese_resultats.xlsx")


@login_required
@csrd_required
def synthese_resultat_IA_par_ESRS_xlsx(request, csrd_id, code_esrs):
    csrd = RapportCSRD.objects.get(id=csrd_id)
    chemin_xlsx = Path(
        settings.BASE_DIR,
        f"reglementations/views/csrd/xlsx/template_synthese_{code_esrs[0]}.xlsx",
    )
    workbook = load_workbook(chemin_xlsx)
    worksheet = workbook[">>>"]
    worksheet["C14"] = normalise_titre_esrs(f"ESRS {code_esrs}")
    worksheet = workbook["Phrases relatives aux ESRS"]
    for document in csrd.documents_analyses:
        _ajoute_ligne_resultat_ia(worksheet, document, True, code_esrs)
    return xlsx_response(workbook, f"resultats_ESRS_{code_esrs}.xlsx")

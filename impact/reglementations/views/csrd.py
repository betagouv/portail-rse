from datetime import datetime
from datetime import timedelta
from functools import wraps
from tempfile import NamedTemporaryFile

import requests
from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist
from django.core.exceptions import PermissionDenied
from django.db.models import Q
from django.http import Http404
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.shortcuts import redirect
from django.template.loader import get_template
from django.template.loader import TemplateDoesNotExist
from django.urls import reverse_lazy
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from openpyxl import Workbook

from entreprises.models import CaracteristiquesAnnuelles
from entreprises.models import Entreprise
from entreprises.views import get_current_entreprise
from habilitations.models import is_user_attached_to_entreprise
from reglementations.enums import ESRS
from reglementations.enums import EtapeCSRD
from reglementations.enums import ETAPES_CSRD
from reglementations.forms.csrd import DocumentAnalyseIAForm
from reglementations.forms.csrd import LienRapportCSRDForm
from reglementations.models import DocumentAnalyseIA
from reglementations.models import RapportCSRD
from reglementations.views.base import Reglementation
from reglementations.views.base import ReglementationAction
from reglementations.views.base import ReglementationStatus


class CSRDReglementation(Reglementation):
    title = "Rapport de Durabilité - CSRD"
    more_info_url = "https://portail-rse.beta.gouv.fr/fiches-reglementaires/rapport-de-durabilite-csrd/"
    tag = "tag-durabilite"
    summary = "Publier un rapport de durabilité."
    zone = "europe"

    @classmethod
    def est_suffisamment_qualifiee(cls, caracteristiques):
        return (
            caracteristiques.date_cloture_exercice is not None
            and caracteristiques.entreprise.est_cotee is not None
            and caracteristiques.entreprise.est_interet_public is not None
            and caracteristiques.effectif is not None
            and caracteristiques.tranche_bilan is not None
            and caracteristiques.tranche_chiffre_affaires is not None
            and caracteristiques.entreprise.appartient_groupe is not None
            and (
                not caracteristiques.entreprise.appartient_groupe
                or (
                    caracteristiques.entreprise.comptes_consolides is not None
                    and (
                        not caracteristiques.entreprise.comptes_consolides
                        or (
                            caracteristiques.effectif_groupe is not None
                            and caracteristiques.tranche_bilan_consolide is not None
                            and caracteristiques.tranche_chiffre_affaires_consolide
                            is not None
                        )
                    )
                )
            )
        )

    @classmethod
    def est_soumis(cls, caracteristiques):
        super().est_soumis(caracteristiques)
        return bool(cls.est_soumis_a_partir_de_l_exercice(caracteristiques))

    @classmethod
    def est_soumis_a_partir_de_l_exercice(
        cls, caracteristiques: CaracteristiquesAnnuelles
    ) -> int | None:
        """Le premier exercice pour lequel l'entreprise est soumise à cette réglementation est l'exercice ouvert à compter du 1er janvier de l'année renvoyée"""
        if not (
            cls.critere_categorie_juridique_sirene(caracteristiques)
            or caracteristiques.entreprise.est_interet_public
        ):
            return
        if cls.est_grand_groupe(caracteristiques):
            if caracteristiques.entreprise.est_societe_mere:
                if (
                    caracteristiques.entreprise.est_cotee
                    or caracteristiques.entreprise.est_interet_public
                ) and caracteristiques.effectif_groupe in (
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_500_ET_4999,
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_5000_ET_9999,
                    CaracteristiquesAnnuelles.EFFECTIF_10000_ET_PLUS,
                ):
                    return 2024
                else:
                    return 2025
            else:
                if cls.est_grande_entreprise(caracteristiques):
                    if (
                        caracteristiques.entreprise.est_cotee
                        and caracteristiques.effectif
                        in (
                            CaracteristiquesAnnuelles.EFFECTIF_ENTRE_500_ET_4999,
                            CaracteristiquesAnnuelles.EFFECTIF_ENTRE_5000_ET_9999,
                            CaracteristiquesAnnuelles.EFFECTIF_10000_ET_PLUS,
                        )
                    ):
                        return 2024
                    else:
                        return 2025
                elif (
                    caracteristiques.entreprise.est_cotee
                    and cls.est_petite_ou_moyenne_entreprise(caracteristiques)
                ):
                    if caracteristiques.effectif_groupe in (
                        CaracteristiquesAnnuelles.EFFECTIF_ENTRE_500_ET_4999,
                        CaracteristiquesAnnuelles.EFFECTIF_ENTRE_5000_ET_9999,
                        CaracteristiquesAnnuelles.EFFECTIF_10000_ET_PLUS,
                    ):
                        return 2024
                    else:
                        return 2025
        else:
            if caracteristiques.entreprise.est_dans_EEE:
                if caracteristiques.entreprise.est_cotee:
                    if cls.est_grande_entreprise(caracteristiques):
                        if caracteristiques.effectif in (
                            CaracteristiquesAnnuelles.EFFECTIF_ENTRE_500_ET_4999,
                            CaracteristiquesAnnuelles.EFFECTIF_ENTRE_5000_ET_9999,
                            CaracteristiquesAnnuelles.EFFECTIF_10000_ET_PLUS,
                        ):
                            return 2024
                        else:
                            return 2025
                    elif cls.est_petite_ou_moyenne_entreprise(caracteristiques):
                        return 2026
                elif caracteristiques.entreprise.est_interet_public:
                    if cls.est_grande_entreprise(caracteristiques):
                        if caracteristiques.effectif in (
                            CaracteristiquesAnnuelles.EFFECTIF_ENTRE_500_ET_4999,
                            CaracteristiquesAnnuelles.EFFECTIF_ENTRE_5000_ET_9999,
                            CaracteristiquesAnnuelles.EFFECTIF_10000_ET_PLUS,
                        ):
                            return 2024
                        else:
                            return 2025
                elif cls.est_grande_entreprise(caracteristiques):
                    return 2025
            else:
                if cls.est_grande_entreprise(caracteristiques):
                    return 2025
                elif (
                    caracteristiques.tranche_chiffre_affaires
                    == CaracteristiquesAnnuelles.CA_100M_ET_PLUS
                ):
                    return 2028

    @classmethod
    def criteres_remplis(cls, caracteristiques):
        criteres = []
        if (
            caracteristiques.entreprise.est_hors_EEE
            and not caracteristiques.entreprise.appartient_groupe
        ):
            criteres.append("votre siège social est hors EEE")
        else:
            if caracteristiques.entreprise.est_cotee:
                criteres.append("votre société est cotée sur un marché réglementé")
            if caracteristiques.entreprise.est_interet_public:
                criteres.append("votre société est d'intérêt public")
            if caracteristiques.entreprise.est_societe_mere and cls.est_grand_groupe(
                caracteristiques
            ):
                criteres.append("votre société est la société mère d'un groupe")
        if critere := cls.critere_effectif(caracteristiques):
            criteres.append(critere)
        if critere := cls.critere_bilan(caracteristiques):
            criteres.append(critere)
        if critere := cls.critere_CA(caracteristiques):
            criteres.append(critere)
        return criteres

    @classmethod
    def critere_effectif(cls, caracteristiques):
        if (
            caracteristiques.entreprise.est_cotee
            or caracteristiques.entreprise.est_interet_public
        ):
            if cls.est_grande_entreprise(caracteristiques):
                if caracteristiques.effectif in (
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_500_ET_4999,
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_5000_ET_9999,
                    CaracteristiquesAnnuelles.EFFECTIF_10000_ET_PLUS,
                ):
                    return "votre effectif est supérieur à 500 salariés"
                elif caracteristiques.effectif in (
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_250_ET_299,
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_300_ET_499,
                ):
                    return "votre effectif est supérieur à 250 salariés"
            if cls.est_grand_groupe(caracteristiques):
                if caracteristiques.effectif_groupe in (
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_500_ET_4999,
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_5000_ET_9999,
                    CaracteristiquesAnnuelles.EFFECTIF_10000_ET_PLUS,
                ):
                    return "l'effectif du groupe est supérieur à 500 salariés"
                elif (
                    caracteristiques.effectif_groupe
                    == CaracteristiquesAnnuelles.EFFECTIF_ENTRE_250_ET_499
                ):
                    return "l'effectif du groupe est supérieur à 250 salariés"
            if cls.est_petite_ou_moyenne_entreprise(caracteristiques):
                if (
                    caracteristiques.effectif
                    != CaracteristiquesAnnuelles.EFFECTIF_MOINS_DE_10
                ):
                    return "votre effectif est supérieur à 10 salariés"
        else:
            if cls.est_grande_entreprise(caracteristiques):
                if caracteristiques.effectif in (
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_250_ET_299,
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_300_ET_499,
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_500_ET_4999,
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_5000_ET_9999,
                    CaracteristiquesAnnuelles.EFFECTIF_10000_ET_PLUS,
                ):
                    return "votre effectif est supérieur à 250 salariés"
            if caracteristiques.entreprise.est_societe_mere and cls.est_grand_groupe(
                caracteristiques
            ):
                if caracteristiques.effectif_groupe in (
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_250_ET_499,
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_500_ET_4999,
                    CaracteristiquesAnnuelles.EFFECTIF_ENTRE_5000_ET_9999,
                    CaracteristiquesAnnuelles.EFFECTIF_10000_ET_PLUS,
                ):
                    return "l'effectif du groupe est supérieur à 250 salariés"

    @classmethod
    def critere_bilan(cls, caracteristiques):
        if cls.est_grand_groupe(caracteristiques):
            if caracteristiques.tranche_bilan_consolide in (
                CaracteristiquesAnnuelles.BILAN_ENTRE_30M_ET_43M,
                CaracteristiquesAnnuelles.BILAN_ENTRE_43M_ET_100M,
                CaracteristiquesAnnuelles.BILAN_100M_ET_PLUS,
            ):
                return "le bilan du groupe est supérieur à 30M€"
        if caracteristiques.tranche_bilan in (
            CaracteristiquesAnnuelles.BILAN_ENTRE_25M_ET_43M,
            CaracteristiquesAnnuelles.BILAN_ENTRE_43M_ET_100M,
            CaracteristiquesAnnuelles.BILAN_100M_ET_PLUS,
        ):
            if cls.est_grande_entreprise(caracteristiques):
                return "votre bilan est supérieur à 25M€"
            elif cls.est_petite_ou_moyenne_entreprise(caracteristiques):
                return "votre bilan est supérieur à 450k€"
        elif (
            caracteristiques.tranche_bilan
            == CaracteristiquesAnnuelles.BILAN_ENTRE_450K_ET_25M
        ):
            if cls.est_petite_ou_moyenne_entreprise(caracteristiques):
                return "votre bilan est supérieur à 450k€"

    @classmethod
    def critere_CA(cls, caracteristiques):
        if cls.est_grand_groupe(caracteristiques):
            if caracteristiques.tranche_chiffre_affaires_consolide in (
                CaracteristiquesAnnuelles.CA_ENTRE_60M_ET_100M,
                CaracteristiquesAnnuelles.CA_100M_ET_PLUS,
            ):
                return "le chiffre d'affaires du groupe est supérieur à 60M€"
        if caracteristiques.tranche_chiffre_affaires in (
            CaracteristiquesAnnuelles.CA_ENTRE_50M_ET_100M,
            CaracteristiquesAnnuelles.CA_100M_ET_PLUS,
        ):
            if cls.est_grande_entreprise(caracteristiques):
                return "votre chiffre d'affaires est supérieur à 50M€"
            elif cls.est_petite_ou_moyenne_entreprise(caracteristiques):
                return "votre chiffre d'affaires est supérieur à 900k€"
        elif (
            caracteristiques.tranche_chiffre_affaires
            == CaracteristiquesAnnuelles.CA_ENTRE_900K_ET_50M
        ):
            if cls.est_petite_ou_moyenne_entreprise(caracteristiques):
                return "votre chiffre d'affaires est supérieur à 900k€"

    @classmethod
    def critere_categorie_juridique_sirene(cls, caracteristiques):
        return (
            caracteristiques.entreprise.categorie_juridique_sirene == 3120
            or 5100 <= caracteristiques.entreprise.categorie_juridique_sirene <= 6199
            or 6300 <= caracteristiques.entreprise.categorie_juridique_sirene <= 6499
            or 8100 <= caracteristiques.entreprise.categorie_juridique_sirene <= 8299
        )

    @classmethod
    def est_delegable(cls, caracteristiques):
        if caracteristiques.entreprise.est_societe_mere:
            return False
        elif (
            cls.est_grande_entreprise(caracteristiques)
            and caracteristiques.entreprise.est_cotee
        ):
            return False
        return cls.est_grand_groupe(caracteristiques)

    @classmethod
    def calculate_status(
        cls,
        caracteristiques: CaracteristiquesAnnuelles,
        user: settings.AUTH_USER_MODEL,
    ) -> ReglementationStatus:
        if reglementation_status := super().calculate_status(caracteristiques, user):
            return reglementation_status

        try:
            rapport = rapport_csrd(
                entreprise=caracteristiques.entreprise,
                user=user,
                annee=datetime.today().year,
            )
        except ObjectDoesNotExist:
            rapport = None

        if rapport and rapport.etape_validee:
            etape_suivante = EtapeCSRD.id_suivant(rapport.etape_validee)
            label_gestion_csrd = "Reprendre l’actualisation de mon rapport"
        else:
            etape_suivante = "introduction"
            label_gestion_csrd = "Actualiser mon Rapport de Durabilité"

        primary_action = ReglementationAction(
            reverse_lazy(
                "reglementations:gestion_csrd",
                kwargs={
                    "siren": caracteristiques.entreprise.siren,
                    "id_etape": etape_suivante,
                },
            ),
            label_gestion_csrd,
        )
        if annee := cls.est_soumis_a_partir_de_l_exercice(caracteristiques):
            premiere_annee_publication = (
                cls.decale_annee_publication_selon_cloture_exercice_comptable(
                    annee, caracteristiques
                )
            )
            exercice_comptable = (
                f"{annee}"
                if caracteristiques.exercice_comptable_est_annee_civile
                else f"{annee}-{annee + 1}"
            )
            status_detail = f"Vous êtes soumis à cette réglementation à partir de {premiere_annee_publication} sur les données de l'exercice comptable {exercice_comptable}"
            if annee == 2028:
                # Ce cas ne peut arriver que pour les micro et PME sans groupe hors EEE
                conditions = "votre société dont le siège social est hors EEE revêt une forme juridique comparable aux sociétés par actions ou aux sociétés à responsabilité limitée, comptabilise un chiffre d'affaires net dans l'Espace économique européen qui excède 150 millions d'euros à la date de clôture des deux derniers exercices consécutifs, ne contrôle ni n'est contrôlée par une autre société et dispose d'une succursale en France dont le chiffre d'affaires net excède 40 millions d'euros"
                status_detail += f" si {conditions}."
            else:
                criteres = cls.criteres_remplis(caracteristiques)
                justification = ", ".join(criteres[:-1]) + " et " + criteres[-1]
                status_detail += f" car {justification}."
                if cls.est_delegable(caracteristiques):
                    status_detail += (
                        " Vous pouvez déléguer cette obligation à votre société-mère."
                    )
            status_detail += " Vous devez publier le Rapport de Durabilité en même temps que le rapport de gestion."
            return ReglementationStatus(
                status=ReglementationStatus.STATUS_SOUMIS,
                status_detail=status_detail,
                primary_action=primary_action,
                prochaine_echeance=premiere_annee_publication,
            )
        else:
            primary_action.title = "Tester un Rapport de Durabilité"
            return ReglementationStatus(
                status=ReglementationStatus.STATUS_NON_SOUMIS,
                status_detail="Vous n'êtes pas soumis à cette réglementation.",
                primary_action=primary_action,
            )

    @staticmethod
    def decale_annee_publication_selon_cloture_exercice_comptable(
        annee, caracteristiques
    ):
        ouverture_exercice_comptable = (
            caracteristiques.date_cloture_exercice + timedelta(days=1)
        ) + relativedelta(
            year=annee
        )  # remplace l'année de la date d'ouverture de l'exercice comptable par la valeur donnée en paramètre (de manière intelligente pour les années bissextiles)
        cloture_exercice_comptable = (
            ouverture_exercice_comptable + relativedelta(months=+12) - timedelta(days=1)
        )
        return (cloture_exercice_comptable + relativedelta(months=+6)).year

    @classmethod
    def est_microentreprise(cls, caracteristiques: CaracteristiquesAnnuelles):
        nombre_seuils_non_depasses = 0
        if (
            caracteristiques.tranche_bilan
            == CaracteristiquesAnnuelles.BILAN_MOINS_DE_450K
        ):
            nombre_seuils_non_depasses += 1
        if (
            caracteristiques.tranche_chiffre_affaires
            == CaracteristiquesAnnuelles.CA_MOINS_DE_900K
        ):
            nombre_seuils_non_depasses += 1
        if caracteristiques.effectif == CaracteristiquesAnnuelles.EFFECTIF_MOINS_DE_10:
            nombre_seuils_non_depasses += 1
        return nombre_seuils_non_depasses >= 2

    @classmethod
    def est_grande_entreprise(cls, caracteristiques: CaracteristiquesAnnuelles) -> bool:
        nombre_seuils_depasses = 0
        if caracteristiques.tranche_bilan in (
            CaracteristiquesAnnuelles.BILAN_ENTRE_25M_ET_43M,
            CaracteristiquesAnnuelles.BILAN_ENTRE_43M_ET_100M,
            CaracteristiquesAnnuelles.BILAN_100M_ET_PLUS,
        ):
            nombre_seuils_depasses += 1
        if caracteristiques.tranche_chiffre_affaires in (
            CaracteristiquesAnnuelles.CA_ENTRE_50M_ET_100M,
            CaracteristiquesAnnuelles.CA_100M_ET_PLUS,
        ):
            nombre_seuils_depasses += 1
        if caracteristiques.effectif in (
            CaracteristiquesAnnuelles.EFFECTIF_ENTRE_250_ET_299,
            CaracteristiquesAnnuelles.EFFECTIF_ENTRE_300_ET_499,
            CaracteristiquesAnnuelles.EFFECTIF_ENTRE_500_ET_4999,
            CaracteristiquesAnnuelles.EFFECTIF_ENTRE_5000_ET_9999,
            CaracteristiquesAnnuelles.EFFECTIF_10000_ET_PLUS,
        ):
            nombre_seuils_depasses += 1
        return nombre_seuils_depasses >= 2

    @classmethod
    def est_petite_ou_moyenne_entreprise(
        cls, caracteristiques: CaracteristiquesAnnuelles
    ) -> bool:
        return not cls.est_microentreprise(
            caracteristiques
        ) and not cls.est_grande_entreprise(caracteristiques)

    @classmethod
    def est_grand_groupe(cls, caracteristiques: CaracteristiquesAnnuelles) -> bool:
        nombre_seuils_depasses = 0
        if caracteristiques.tranche_bilan_consolide in (
            CaracteristiquesAnnuelles.BILAN_ENTRE_30M_ET_43M,
            CaracteristiquesAnnuelles.BILAN_ENTRE_43M_ET_100M,
            CaracteristiquesAnnuelles.BILAN_100M_ET_PLUS,
        ):
            nombre_seuils_depasses += 1
        if caracteristiques.tranche_chiffre_affaires_consolide in (
            CaracteristiquesAnnuelles.CA_ENTRE_60M_ET_100M,
            CaracteristiquesAnnuelles.CA_100M_ET_PLUS,
        ):
            nombre_seuils_depasses += 1
        if caracteristiques.effectif_groupe in (
            CaracteristiquesAnnuelles.EFFECTIF_ENTRE_250_ET_499,
            CaracteristiquesAnnuelles.EFFECTIF_ENTRE_500_ET_4999,
            CaracteristiquesAnnuelles.EFFECTIF_ENTRE_5000_ET_9999,
            CaracteristiquesAnnuelles.EFFECTIF_10000_ET_PLUS,
        ):
            nombre_seuils_depasses += 1
        return (
            caracteristiques.entreprise.appartient_groupe
            and nombre_seuils_depasses >= 2
        )


def rapport_csrd(user, entreprise, annee):
    """Cherche un RapportCSRD

    Lève une exception si le RapportCSRD ou l'Habilitation n'existe pas
    """
    habilitation = user.habilitation_set.get(entreprise=entreprise)
    return RapportCSRD.objects.get(
        entreprise=entreprise,
        proprietaire=None if habilitation.is_confirmed else user,
        annee=annee,
    )


@login_required
def guide_csrd(request, siren=None, phase=0, etape=0, sous_etape=0):
    if not siren:
        entreprise = get_current_entreprise(request)
        if not entreprise:
            messages.warning(
                request,
                "Commencez par ajouter une entreprise à votre compte utilisateur avant d'accéder à l'espace Rapport de Durabilité",
            )
            return redirect("entreprises:entreprises")
        return redirect("reglementations:csrd", siren=entreprise.siren)

    entreprise = get_object_or_404(Entreprise, siren=siren)
    if not is_user_attached_to_entreprise(request.user, entreprise):
        raise PermissionDenied

    if sous_etape:
        template_name = (
            f"reglementations/espace_csrd/phase{phase}_etape{etape}_{sous_etape}.html"
        )
    elif etape:
        template_name = f"reglementations/espace_csrd/phase{phase}_etape{etape}.html"
    elif phase:
        template_name = f"reglementations/espace_csrd/phase{phase}.html"
    else:
        template_name = "reglementations/espace_csrd/index.html"

    try:
        template = get_template(template_name)
    except TemplateDoesNotExist:
        raise Http404

    context = {
        "entreprise": entreprise,
        "phase": phase,
        "etape": etape,
        "sous_etape": sous_etape,
    }

    return HttpResponse(template.render(context, request))


@login_required
def gestion_csrd(request, siren=None, id_etape="introduction"):
    if not siren:
        entreprise = get_current_entreprise(request)
        if not entreprise:
            messages.warning(
                request,
                "Commencez par ajouter une entreprise à votre compte utilisateur avant d'accéder à l'espace Rapport de Durabilité",
            )
            return redirect("entreprises:entreprises")
        return redirect("reglementations:csrd", siren=entreprise.siren)

    entreprise = get_object_or_404(Entreprise, siren=siren)
    if not is_user_attached_to_entreprise(request.user, entreprise):
        raise PermissionDenied

    template_name = f"reglementations/csrd/etape-{id_etape}.html"
    try:
        template = get_template(template_name)
    except TemplateDoesNotExist:
        raise Http404

    # En analysant un peu les requêtes exécutées,
    # les fetch sur les habilitations et l'entreprise se répètent (requêtes identiques)
    # un peu de centralisation à faire pour éviter les répétitions dans `utils.middlewares.ExtendUserMiddleware`

    # par ex., on peut récupérer l'habilitation pour cette entreprise, elle est déjà en cache
    habilitation = request.user.habilitation_set.get(entreprise=entreprise)
    annee = datetime.now().year

    if request.method == "POST":
        try:
            csrd = rapport_csrd(request.user, entreprise, annee)
        except ObjectDoesNotExist:
            raise Http404
        csrd.etape_validee = id_etape
        csrd.save()
        return redirect(
            "reglementations:gestion_csrd",
            siren=siren,
            id_etape=EtapeCSRD.id_suivant(id_etape),
        )

    # on vérifie si on a sélectionné une année manuellement
    if csrd_id := request.session.get("rapport_csrd_courant"):
        try:
            csrd = RapportCSRD.objects.prefetch_related("enjeux", "enjeux__parent").get(
                pk=csrd_id, entreprise=entreprise
            )
        except RapportCSRD.DoesNotExist:
            # par ex. : l'utilisateur a selectionné une autre entreprise
            request.session.pop("rapport_csrd_courant")

    if not request.session.get("rapport_csrd_courant"):
        # les prefetch de l'enjeu parent évitent des N+1 au niveau du template
        csrd, _ = RapportCSRD.objects.prefetch_related(
            "enjeux", "enjeux__parent"
        ).get_or_create(
            entreprise=entreprise,
            proprietaire=None if habilitation.is_confirmed else request.user,
            annee=annee,
        )

    # on récupère les rapports si plusieurs existants, pour permettre une sélection
    rapports_csrd = RapportCSRD.objects.filter(
        entreprise=entreprise,
        proprietaire=None if habilitation.is_confirmed else request.user,
    ).order_by("annee")

    context = {
        "entreprise": entreprise,
        "etape": EtapeCSRD.get(id_etape),
        "csrd": csrd,
        "annee": annee,
        "steps": ETAPES_CSRD,
        "rapports_csrd": rapports_csrd,
    }

    match EtapeCSRD.get(id_etape).id:
        ## légèrement plus lisible qu'un `if`
        case "selection-informations":
            nb_enjeux_non_analyses = csrd.enjeux.non_analyses().count()
            context |= {
                "can_download": nb_enjeux_non_analyses
                != csrd.enjeux.selectionnes().count(),
                "nb_enjeux_non_analyses": nb_enjeux_non_analyses,
            }
        case "analyse-ecart":
            context |= {
                "form": DocumentAnalyseIAForm(),
                "documents": csrd.documents,
            }
        case "redaction-rapport-durabilite":
            context |= {"form": LienRapportCSRDForm(instance=csrd)}

    return HttpResponse(template.render(context, request))


def csrd_required(function):
    @wraps(function)
    def wrap(request, siren):
        entreprise = get_object_or_404(Entreprise, siren=siren)
        try:
            habilitation = request.user.habilitation_set.get(entreprise=entreprise)
            csrd = RapportCSRD.objects.prefetch_related("enjeux", "enjeux__parent").get(
                entreprise=entreprise,
                proprietaire=None if habilitation.is_confirmed else request.user,
                annee=datetime.now().year,
            )
            return function(request, siren, csrd=csrd)
        except ObjectDoesNotExist:
            raise Http404("Ce rapport de durabilité n'existe pas")

    return wrap


@login_required
@csrd_required
def enjeux_xlsx(request, siren, csrd=None):
    return _build_xlsx(csrd.enjeux, csrd=csrd)


@login_required
@csrd_required
def enjeux_materiels_xlsx(request, siren, csrd=None):
    return _build_xlsx(csrd.enjeux, csrd=csrd, materiels=True)


def _build_xlsx(enjeux, csrd=None, materiels=False):
    # construit le fichier XLS pour :
    # - les enjeux sélectionnés
    # - les enjeux analysés (matériels ou non)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "ESRS"
    worksheet["B1"] = "Enjeux de durabilité"
    worksheet["C1"] = "Descriptif"
    worksheet["D1"] = "Origine"

    if materiels:
        worksheet["E1"] = "Materialité"

    numero_ligne = 2
    for esrs in ESRS:
        enjeux = csrd.enjeux_par_esrs(esrs)

        if materiels:
            enjeux = enjeux.filter(materiel__isnull=False)

        for enjeu in enjeux:
            if enjeu.selection:
                worksheet[f"A{numero_ligne}"] = enjeu.esrs
                worksheet[f"B{numero_ligne}"] = enjeu.nom
                worksheet[f"C{numero_ligne}"] = enjeu.description
                worksheet[f"D{numero_ligne}"] = (
                    "personnel" if enjeu.modifiable else "AR"
                )

                if materiels:
                    worksheet[f"E{numero_ligne}"] = (
                        "Matériel" if enjeu.materiel else "Non-matériel"
                    )

                numero_ligne += 1

    filename = "enjeux_csrd.xlsx" if not materiels else "enjeux_csrd_materiels.xlsx"
    return _xlsx_response(workbook, filename)


def _xlsx_response(workbook, filename):
    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        xlsx_stream = tmp.read()

    response = HttpResponse(
        xlsx_stream,
        content_type="application/vnd.openxmlformatsofficedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"filename={filename}"

    return response


@login_required
@csrd_required
def datapoints_xlsx(request, _, csrd=None):
    materiel = request.GET.get("materiel", True) != "false"
    esrs_a_supprimer = _esrs_materiel_a_supprimer(csrd, materiel)
    workbook = load_workbook("impact/static/CSRD/ESRS_Data_Points_EFRAG.xlsx")
    for esrs in esrs_a_supprimer:
        titre_onglet = esrs.replace("_", " ")
        workbook.remove(workbook[titre_onglet])

    filename = (
        "datapoints_csrd_materiels.xlsx"
        if materiel
        else "datapoints_csrd_non_materiels.xlsx"
    )
    return _xlsx_response(workbook, filename)


def _esrs_materiel_a_supprimer(csrd: RapportCSRD, materiel: bool):
    tous_les_esrs = set(ESRS.values)

    if materiel:
        enjeux_materiels = csrd.enjeux.materiels()
        esrs_materiels = set((enjeu.esrs for enjeu in enjeux_materiels))
        esrs_a_supprimer = tous_les_esrs - esrs_materiels
        esrs_a_supprimer.discard("ESRS_2")
    else:
        # les enjeux non-analysés doivent aussi être dans ce document (comme non-matériels)
        enjeux_non_materiels = csrd.enjeux.filter(
            Q(materiel=False) | Q(selection=False)
        )
        esrs_non_materiels = set((enjeu.esrs for enjeu in enjeux_non_materiels))
        # pour les enjeux non-matériels, ESRS_2 et ESRS_2_MDR sont à supprimer
        esrs_a_supprimer = tous_les_esrs - esrs_non_materiels
        esrs_a_supprimer |= {"ESRS2_MDR", "ESRS_2"}

    # ne pas supprimer ESRS_1 car il n'existe pas dans le fichier .xlsx
    esrs_a_supprimer.discard("ESRS_1")

    return esrs_a_supprimer


def lance_analyse_IA(request, id_document):
    url = f"{settings.IA_BASE_URL}/run-task"
    try:
        document = DocumentAnalyseIA.objects.get(id=id_document)
    except ObjectDoesNotExist:
        raise Http404("Ce document n'existe pas")

    if document.etat != "success":
        response = requests.post(
            url, {"document_id": document.id, "url": document.fichier.url}
        )
        document.etat = response.json()["status"]
        document.save()

    referer = request.META.get("HTTP_REFERER")
    return redirect(referer)


@csrf_exempt
def resultat_analyse_IA(request, id_document):
    try:
        document = DocumentAnalyseIA.objects.get(id=id_document)
    except ObjectDoesNotExist:
        raise Http404("Ce document n'existe pas")

    status = request.POST.get("status")
    document.etat = status
    if status == "success":
        document.resultat_csv = request.POST["resultat_csv"]
    document.save()

    return HttpResponse("OK")


def csv_analyse_IA(request, id_document):
    try:
        document = DocumentAnalyseIA.objects.get(id=id_document)
    except ObjectDoesNotExist:
        raise Http404("Ce document n'existe pas")

    response = HttpResponse(
        document.resultat_csv,
        content_type="text/csv",
    )
    response["Content-Disposition"] = f"filename=resultats_analyse_ia.csv"
    return response


def synthese_resultat_IA_xlsx(request, id_csrd):
    try:
        csrd = RapportCSRD.objects.get(id=id_csrd)
    except ObjectDoesNotExist:
        raise Http404("Ce document n'existe pas")

    workbook = Workbook()
    worksheet = workbook.active
    return _xlsx_response(workbook, "synthese_resultats.xlsx")
